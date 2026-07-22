#!/usr/bin/env python3
# SPDX-License-Identifier: Apache-2.0
"""End-to-end smoke test for the public XLSX command-line interface."""

from __future__ import annotations

import argparse
import hashlib
import importlib.util
import json
import shutil
import subprocess
import sys
import tempfile
import time
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

TOOL = Path(__file__).with_name("xlsx_tool.py")
SUBPROCESS_TIMEOUT_SECONDS = 120


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def run_tool(
    *arguments: str | Path,
    expect_status: int = 0,
    timeout_seconds: float = SUBPROCESS_TIMEOUT_SECONDS,
) -> dict[str, Any]:
    command = [
        sys.executable,
        str(TOOL),
        "--debug",
        *(str(argument) for argument in arguments),
    ]
    try:
        completed = subprocess.run(
            command,
            check=False,
            capture_output=True,
            text=True,
            encoding="utf-8",
            timeout=timeout_seconds,
        )
    except subprocess.TimeoutExpired as exc:
        stdout = (
            exc.stdout.decode("utf-8", "replace") if isinstance(exc.stdout, bytes) else exc.stdout
        )
        stderr = (
            exc.stderr.decode("utf-8", "replace") if isinstance(exc.stderr, bytes) else exc.stderr
        )
        raise AssertionError(
            f"CLI subprocess timed out after {timeout_seconds:g}s: {command!r}\n"
            f"stdout={stdout or ''}\nstderr={stderr or ''}"
        ) from exc
    if completed.returncode != expect_status:
        raise AssertionError(
            f"Unexpected status {completed.returncode}, expected {expect_status}.\n"
            f"stdout={completed.stdout}\nstderr={completed.stderr}"
        )
    stream = completed.stdout if expect_status == 0 else completed.stderr
    try:
        payload = json.loads(stream)
    except json.JSONDecodeError as exc:
        raise AssertionError(
            f"CLI did not emit one JSON document.\nstdout={completed.stdout}\n"
            f"stderr={completed.stderr}"
        ) from exc
    if payload.get("ok") is not (expect_status == 0):
        raise AssertionError(f"Unexpected JSON status: {payload}")
    return payload


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


def create_fixture(temp: Path) -> Path:
    job = {
        "schema_version": 1,
        "properties": {"title": "Quarterly café sales", "creator": "xlsx smoke test"},
        "named_styles": [
            {
                "name": "ReportHeader",
                "font": {"bold": True, "color": "FFFFFF"},
                "fill": {"color": "1F4E78"},
                "alignment": {"horizontal": "center"},
            }
        ],
        "sheets": [
            {
                "name": "Sales",
                "rows": [
                    ["Region", "Product", "Revenue", "Units", "Code", "Untrusted"],
                    ["North", "Café", 120.5, 3, "00123", "=2+2"],
                    ["South", "Thé", 80, 2, "00007", "+SUM(A1:A2)"],
                    ["North", "Crème", 150, 5, "00999", "ordinary"],
                    ["South", "Café", 100, 4, "00420", "@command"],
                ],
                "cells": {
                    "A1": {"value": "Region", "style": "ReportHeader"},
                    "B1": {"value": "Product", "style": "ReportHeader"},
                    "C1": {"value": "Revenue", "style": "ReportHeader"},
                    "D1": {"value": "Units", "style": "ReportHeader"},
                    "E1": {"value": "Code", "style": "ReportHeader"},
                    "F1": {"value": "Untrusted", "style": "ReportHeader"},
                    "H1": {"value": "Average revenue", "style": "ReportHeader"},
                    "H2": {"formula": "=AVERAGE(C2:C5)", "number_format": "$0.00"},
                },
                "freeze_panes": "A2",
                "dimensions": {"columns": {"A": 14, "B": 16, "C": 13, "F": 18}},
                "tables": [
                    {
                        "name": "SalesTable",
                        "range": "A1:F5",
                        "style": {"name": "TableStyleMedium2"},
                    }
                ],
                "conditional_formats": [
                    {
                        "range": "C2:C5",
                        "type": "color_scale",
                        "colors": ["F8696B", "FFEB84", "63BE7B"],
                    }
                ],
                "data_validations": [
                    {
                        "range": "D2:D100",
                        "type": "whole",
                        "operator": "greaterThan",
                        "formula1": "0",
                    }
                ],
                "charts": [
                    {
                        "type": "bar",
                        "title": "Revenue by row",
                        "data": "C1:C5",
                        "categories": "B2:B5",
                        "anchor": "J2",
                    },
                    {
                        "type": "scatter",
                        "title": "Revenue vs units",
                        "x_values": "D2:D5",
                        "y_values": ["C2:C5"],
                        "series_titles": ["Revenue"],
                        "anchor": "J20",
                    },
                ],
            },
            {
                "name": "Notes",
                "rows": [
                    ["Status", "Owner", "Dependency"],
                    ["Draft", "Zoë", {"formula": "=sales!A1"}],
                ],
                "state": "visible",
            },
            {"name": "Blank", "rows": []},
            {
                "name": "Raw Headers",
                "rows": [["", "", "=Header"], ["first", "second", "ordinary"]],
            },
        ],
        "defined_names": [{"name": "RevenueData", "refers_to": "'Sales'!$C$2:$C$5"}],
        "calculation": {"mode": "auto"},
    }
    job_path = temp / "create.json"
    output = temp / "created.xlsx"
    write_json(job_path, job)
    payload = run_tool("create", job_path, output)
    assert payload["result"]["verification"]["reopened"]
    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == ["Sales", "Notes", "Blank", "Raw Headers"]
    assert workbook["Sales"]["H2"].data_type == "f"
    assert workbook["Sales"]["F2"].data_type == "s"
    assert len(workbook["Sales"].tables) == 1
    assert len(workbook["Sales"].conditional_formatting) == 1
    assert len(workbook["Sales"]._charts) == 2
    scatter = workbook["Sales"]._charts[1]
    assert type(scatter).__name__ == "ScatterChart"
    assert len(scatter.series) == 1
    assert scatter.scatterStyle is not None
    workbook.close()
    return output


def inspect_extract_edit(temp: Path, source: Path) -> Path:
    before = sha256(source)
    inspection = run_tool("inspect", source, "--max-cells", "100")
    assert inspection["result"]["counts"]["sheets"] == 4
    assert inspection["result"]["counts"]["charts"] == 2
    assert inspection["result"]["counts"]["formulas"] == 2

    extracted = temp / "sales.json"
    extraction = run_tool(
        "extract",
        source,
        extracted,
        "--sheet",
        "Sales",
        "--range",
        "A1:H2",
        "--format",
        "json",
    )
    assert extraction["result"]["rows"] == 2
    extracted_payload = json.loads(extracted.read_text(encoding="utf-8"))
    assert extracted_payload["rows"][1][5] == "=2+2"
    assert extracted_payload["rows"][1][7] == "=AVERAGE(C2:C5)"

    job = {
        "schema_version": 1,
        "operations": [
            {"op": "set_cell", "sheet": "Sales", "cell": "B2", "value": "Café au lait"},
            {
                "op": "set_range",
                "sheet": "Sales",
                "range": "I1:I2",
                "values": [["Review"], [{"value": "Ready", "font": {"bold": True}}]],
            },
            {
                "op": "format_range",
                "sheet": "Sales",
                "range": "C2:C5",
                "style": {"number_format": "$#,##0.00"},
            },
            {
                "op": "update_table",
                "table": {
                    "name": "SalesTable",
                    "style": {"name": "TableStyleMedium9", "show_row_stripes": True},
                    "totals_row_shown": True,
                },
            },
            {
                "op": "update_chart",
                "sheet": "Sales",
                "index": 0,
                "chart": {
                    "type": "line",
                    "title": "Updated revenue",
                    "data": "C1:C5",
                    "categories": "B2:B5",
                },
            },
            {
                "op": "add_sheet",
                "name": "Audit",
                "rows": [["Action", "Count"], ["Updated", 1]],
            },
        ],
    }
    job_path = temp / "edit.json"
    output = temp / "edited.xlsx"
    write_json(job_path, job)
    edit = run_tool("edit", source, job_path, output)
    assert edit["result"]["operations"] == 6
    assert any("rewrites the OOXML package" in warning for warning in edit["warnings"])
    assert sha256(source) == before
    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == ["Sales", "Notes", "Blank", "Raw Headers", "Audit"]
    assert workbook["Sales"]["B2"].value == "Café au lait"
    assert workbook["Sales"]["C2"].number_format == "$#,##0.00"
    assert workbook["Sales"].tables["SalesTable"].totalsRowShown is True
    assert len(workbook["Sales"]._charts) == 2
    workbook.close()
    return output


def clean_and_summarize(temp: Path, source: Path) -> Path:
    clean_job = {
        "schema_version": 1,
        "source": {"sheet": "Sales", "range": "A1:F5", "header_row": 1},
        "destination_sheet": "Cleaned",
        "policies": {
            "whitespace": {"strip": True, "collapse": True, "columns": ["Product"]},
            "coercion": {
                "Revenue": {"type": "float", "errors": "raise"},
                "Units": {"type": "integer", "errors": "raise"},
            },
            "duplicates": {"subset": ["Region", "Product"], "keep": "first"},
            "sort": [{"column": "Revenue", "ascending": False}],
        },
        "table": {"name": "CleanedSales"},
    }
    clean_job_path = temp / "clean.json"
    cleaned = temp / "cleaned.xlsx"
    write_json(clean_job_path, clean_job)
    clean = run_tool("clean", source, clean_job_path, cleaned)
    assert clean["result"]["counts"]["output_rows"] == 4
    workbook = load_workbook(cleaned)
    assert "Sales" in workbook.sheetnames and "Cleaned" in workbook.sheetnames
    assert workbook["Cleaned"]["C2"].value == 150
    assert workbook["Cleaned"].auto_filter.ref is None
    assert "CleanedSales" in workbook["Cleaned"].tables
    workbook.close()

    summary_job = {
        "schema_version": 1,
        "source": {"sheet": "Sales", "range": "A1:F5", "header_row": 1},
        "destination_sheet": "Static Summary",
        "group_by": ["Region"],
        "values": [
            {"column": "Revenue", "aggregation": "sum", "name": "Revenue total"},
            {"column": "Units", "aggregation": "sum", "name": "Unit total"},
        ],
        "chart": {"type": "bar", "title": "Revenue and units by region"},
    }
    summary_job_path = temp / "summary.json"
    summarized = temp / "summarized.xlsx"
    write_json(summary_job_path, summary_job)
    summary = run_tool("summarize", source, summary_job_path, summarized)
    assert summary["result"]["rows"] == 2
    workbook = load_workbook(summarized)
    sheet = workbook["Static Summary"]
    assert sheet["A1"].value.startswith("STATIC SUMMARY")
    assert len(sheet._charts) == 1
    workbook.close()
    return summarized


def round_trip_delimited(temp: Path) -> None:
    csv_source = temp / "utf8.csv"
    csv_source.write_text(
        'Code,City,Text,Amount\n00123,Montréal,"café, crème",10.5\n00007,Zürich,"=2+2",20\n',
        encoding="utf-8",
    )
    schema_path = temp / "schema.json"
    write_json(schema_path, {"Code": "string", "Amount": "float"})
    workbook_path = temp / "from-csv.xlsx"
    conversion = run_tool(
        "convert",
        csv_source,
        workbook_path,
        "--schema",
        schema_path,
        "--leading-zeros",
        "preserve",
    )
    assert conversion["result"]["formula_injection_guarded"] == 1
    workbook = load_workbook(workbook_path, data_only=False)
    assert workbook["Data"]["A2"].value == "00123"
    assert workbook["Data"]["C3"].value == "=2+2"
    assert workbook["Data"]["C3"].data_type == "s"
    workbook.close()

    no_header_source = temp / "no-header.csv"
    no_header_source.write_text("00123,NA,\n00007,,value\n", encoding="utf-8")
    no_header_workbook = temp / "no-header.xlsx"
    no_header = run_tool(
        "convert",
        no_header_source,
        no_header_workbook,
        "--no-header",
        "--input-na-policy",
        "literal",
    )
    assert no_header["result"]["header_policy"] == "none"
    workbook = load_workbook(no_header_workbook)
    assert workbook["Data"]["A1"].value == "00123"
    assert workbook["Data"]["B1"].value == "NA"
    assert workbook["Data"]["A3"].value is None
    workbook.close()

    custom_na_workbook = temp / "custom-na.xlsx"
    custom_na = run_tool(
        "convert",
        csv_source,
        custom_na_workbook,
        "--input-na-policy",
        "custom",
        "--input-na-value",
        "20",
    )
    assert custom_na["result"]["input_na_policy"] == "custom"
    workbook = load_workbook(custom_na_workbook)
    assert workbook["Data"]["D3"].value is None
    workbook.close()

    tsv_output = temp / "roundtrip.tsv"
    exported = run_tool(
        "convert",
        workbook_path,
        tsv_output,
        "--output-format",
        "tsv",
        "--sheet",
        "Data",
    )
    assert exported["result"]["formula_injection_guarded"] == 1
    text = tsv_output.read_text(encoding="utf-8")
    assert "Montréal" in text and "00123" in text and "'=2+2" in text

    second_workbook = temp / "from-tsv.xlsx"
    run_tool(
        "convert",
        tsv_output,
        second_workbook,
        "--input-format",
        "tsv",
        "--leading-zeros",
        "preserve",
    )
    workbook = load_workbook(second_workbook)
    assert workbook["Data"]["A2"].value == "00123"
    assert workbook["Data"]["B3"].value == "Zürich"
    workbook.close()

    exported_directory = temp / "all-sheets"
    all_sheets = run_tool(
        "convert",
        second_workbook,
        exported_directory,
        "--all-sheets",
        "--output-format",
        "csv",
    )
    assert all_sheets["result"]["verification"]["file_count"] == 1
    assert (exported_directory / "001-Data.csv").is_file()

    header_source = temp / "header-collision.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["=X", "'=X"])
    worksheet.append(["=1+1", "literal"])
    workbook.save(header_source)
    workbook.close()
    header_output = temp / "header-collision.csv"
    header_export = run_tool(
        "convert",
        header_source,
        header_output,
        "--sheet-policy",
        "header",
        "--values",
        "formulas",
    )
    assert header_export["result"]["sheet_policy"] == "header"
    assert header_export["result"]["formula_injection_guarded"] == 2
    assert header_output.read_text(encoding="utf-8").splitlines() == [
        "'=X,'=X",
        "'=1+1,literal",
    ]

    failure = run_tool(
        "convert",
        header_source,
        temp / "ignored-xlsx-option.csv",
        "--schema",
        schema_path,
        expect_status=2,
    )
    assert failure["error"]["category"] == "bad_input"
    failure = run_tool(
        "convert",
        csv_source,
        temp / "ignored-csv-option.xlsx",
        "--range",
        "A1:B2",
        expect_status=2,
    )
    assert failure["error"]["category"] == "bad_input"
    failure = run_tool(
        "convert",
        header_source,
        temp / "explicit-xlsx-leading-zero-option.csv",
        "--leading-zeros",
        "preserve",
        expect_status=2,
    )
    assert failure["error"]["category"] == "bad_input"
    failure = run_tool(
        "convert",
        csv_source,
        temp / "explicit-csv-sheet-policy.xlsx",
        "--sheet-policy",
        "raw",
        expect_status=2,
    )
    assert failure["error"]["category"] == "bad_input"


def raw_all_sheet_export(temp: Path, source: Path) -> None:
    exported_directory = temp / "rich-all-sheets"
    payload = run_tool(
        "convert",
        source,
        exported_directory,
        "--all-sheets",
        "--output-format",
        "csv",
    )
    files = payload["result"]["files"]
    assert payload["result"]["sheet_policy"] == "raw"
    assert payload["result"]["verification"]["file_count"] == 6
    assert files[2]["sheet"] == "Blank" and files[2]["rows"] == 0
    assert (exported_directory / "003-Blank.csv").read_text(encoding="utf-8") == ""
    raw_headers = (exported_directory / "004-Raw_Headers.csv").read_text(encoding="utf-8")
    assert raw_headers.startswith(",,'=Header\n")
    static_summary = (exported_directory / "006-Static_Summary.csv").read_text(encoding="utf-8")
    assert static_summary.startswith("STATIC SUMMARY")


def case_insensitive_dependency(temp: Path, source: Path) -> None:
    job_path = temp / "case-dependency.json"
    write_json(
        job_path,
        {
            "schema_version": 1,
            "operations": [{"op": "rename_sheet", "sheet": "Sales", "new_name": "Renamed Sales"}],
        },
    )
    failure = run_tool("edit", source, job_path, temp / "case-dependency.xlsx", expect_status=5)
    assert failure["error"]["category"] == "ambiguous_edit"
    assert failure["error"]["details"]["dependency_count"] >= 1


def failure_path(temp: Path, source: Path) -> None:
    macro_path = temp / "refused.xlsm"
    macro_path.write_bytes(b"PK\x03\x04not-a-real-workbook")
    failure = run_tool("inspect", macro_path, expect_status=3)
    assert failure["error"]["category"] == "unsupported_operation"

    macro_package = temp / "macro-bearing.xlsx"
    with zipfile.ZipFile(macro_package, "w") as package:
        package.writestr(
            "[Content_Types].xml",
            '<?xml version="1.0"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Override PartName="/xl/vbaProject.bin" '
            'ContentType="application/vnd.ms-office.vbaProject"/>'
            "</Types>",
        )
        package.writestr("xl/workbook.xml", "<workbook/>")
        package.writestr("xl/vbaProject.bin", b"not executable")
    failure = run_tool("inspect", macro_package, expect_status=3)
    assert failure["error"]["category"] == "unsupported_operation"

    dtd_package = temp / "utf16-dtd.xlsx"
    with zipfile.ZipFile(dtd_package, "w") as package:
        package.writestr(
            "[Content_Types].xml",
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"/>',
        )
        package.writestr(
            "xl/workbook.xml",
            (
                '<?xml version="1.0" encoding="UTF-16"?>'
                '<!DOCTYPE workbook [<!ENTITY probe "blocked">]><workbook>&probe;</workbook>'
            ).encode("utf-16"),
        )
    failure = run_tool("inspect", dtd_package, expect_status=3)
    assert failure["error"]["category"] == "unsupported_operation"

    invalid_encoding_package = temp / "invalid-encoding.xlsx"
    with zipfile.ZipFile(invalid_encoding_package, "w") as package:
        package.writestr(
            "[Content_Types].xml",
            '<?xml version="1.0"?><Types '
            'xmlns="http://schemas.openxmlformats.org/package/2006/content-types"/>',
        )
        package.writestr(
            "xl/workbook.xml",
            "<?xml version='1.0' encoding='x-not-real'?><workbook/>",
        )
    failure = run_tool("inspect", invalid_encoding_package, expect_status=2)
    assert failure["error"]["category"] == "bad_input"

    bad_encoding_output = temp / "bad-encoding.csv"
    failure = run_tool(
        "extract",
        source,
        bad_encoding_output,
        "--format",
        "csv",
        "--encoding",
        "x-not-real",
        expect_status=2,
    )
    assert failure["error"]["category"] == "bad_input"
    assert not bad_encoding_output.exists()

    invalid_sheet_job = temp / "invalid-sheet.json"
    write_json(
        invalid_sheet_job,
        {"schema_version": 1, "sheets": [{"name": "Bad/Name", "rows": []}]},
    )
    failure = run_tool(
        "create",
        invalid_sheet_job,
        temp / "invalid-sheet.xlsx",
        expect_status=2,
    )
    assert failure["error"]["category"] == "bad_input"

    invalid_style_job = temp / "invalid-style.json"
    write_json(
        invalid_style_job,
        {
            "schema_version": 1,
            "sheets": [
                {
                    "name": "Data",
                    "cells": {"A1": {"value": "x", "font": {"underline": "invalid"}}},
                }
            ],
        },
    )
    failure = run_tool(
        "create",
        invalid_style_job,
        temp / "invalid-style.xlsx",
        expect_status=2,
    )
    assert failure["error"]["category"] == "bad_input"

    unknown_key_job = temp / "unknown-key.json"
    write_json(
        unknown_key_job,
        {
            "schema_version": 1,
            "operations": [
                {
                    "op": "set_cell",
                    "sheet": "Sales",
                    "cell": "A2",
                    "value": "x",
                    "vale": "typo",
                }
            ],
        },
    )
    failure = run_tool(
        "edit",
        source,
        unknown_key_job,
        temp / "unknown-key.xlsx",
        expect_status=2,
    )
    assert failure["error"]["category"] == "bad_input"

    structural_unknown_key_job = temp / "structural-unknown-key.json"
    write_json(
        structural_unknown_key_job,
        {
            "schema_version": 1,
            "operations": [
                {
                    "op": "insert_rows",
                    "sheet": "Sales",
                    "index": 2,
                    "amount": 1,
                    "typo": True,
                }
            ],
        },
    )
    failure = run_tool(
        "edit",
        source,
        structural_unknown_key_job,
        temp / "structural-unknown-key.xlsx",
        expect_status=2,
    )
    assert failure["error"]["category"] == "bad_input"

    invalid_table_job = temp / "invalid-table.json"
    write_json(
        invalid_table_job,
        {
            "schema_version": 1,
            "operations": [
                {
                    "op": "update_table",
                    "table": {"name": "SalesTable", "new_name": "1Invalid"},
                }
            ],
        },
    )
    failure = run_tool(
        "edit",
        source,
        invalid_table_job,
        temp / "invalid-table.xlsx",
        expect_status=2,
    )
    assert failure["error"]["category"] == "bad_input"


def external_link_warning(temp: Path) -> None:
    from openpyxl.packaging.relationship import Relationship
    from openpyxl.workbook.external_link.external import ExternalBook, ExternalLink

    source = temp / "external-link.xlsx"
    workbook = Workbook()
    link = ExternalLink(externalBook=ExternalBook(id="rId1"))
    link.file_link = Relationship(
        Type="externalLinkPath",
        Target="file:///untrusted/source.xlsx",
        TargetMode="External",
    )
    workbook._external_links = [link]
    workbook.save(source)
    workbook.close()

    job_path = temp / "external-link-edit.json"
    write_json(
        job_path,
        {
            "schema_version": 1,
            "operations": [{"op": "set_cell", "sheet": "Sheet", "cell": "A1", "value": "updated"}],
        },
    )
    payload = run_tool("edit", source, job_path, temp / "external-link-edited.xlsx")
    assert any("external link(s)" in warning for warning in payload["warnings"])

    failure = run_tool("convert", source, temp / "external-link.pdf", expect_status=3)
    assert failure["error"]["category"] == "unsupported_operation"


def merge_and_sheet_settings(temp: Path, source: Path) -> None:
    refused_job = temp / "merge-refused.json"
    write_json(
        refused_job,
        {
            "schema_version": 1,
            "operations": [{"op": "set_sheet", "sheet": "Raw Headers", "merges": ["A2:B2"]}],
        },
    )
    failure = run_tool("edit", source, refused_job, temp / "merge-refused.xlsx", expect_status=5)
    assert failure["error"]["category"] == "ambiguous_edit"
    assert failure["error"]["details"]["occupied_cells"] == ["B2"]

    merged_job = temp / "merge-acknowledged.json"
    write_json(
        merged_job,
        {
            "schema_version": 1,
            "operations": [
                {
                    "op": "set_sheet",
                    "sheet": "Raw Headers",
                    "merges": ["A2:B2"],
                    "allow_merge_data_loss": True,
                    "tab_color": "FF3366CC",
                    "show_gridlines": False,
                    "zoom_scale": 150,
                    "auto_width": {"columns": ["C"], "max_width": 24},
                }
            ],
        },
    )
    merged_path = temp / "merge-acknowledged.xlsx"
    payload = run_tool("edit", source, merged_job, merged_path)
    assert payload["result"]["counts"]["ranges_merged"] == 1
    assert payload["result"]["counts"]["columns_auto_sized"] == 1
    assert any("discarded 1 non-top-left cell value(s)" in w for w in payload["warnings"])
    inspection = run_tool("inspect", merged_path, "--sheet", "Raw Headers", "--no-cells")
    sheet = inspection["result"]["sheets"][0]
    assert sheet["merged_cells"] == ["A2:B2"]
    assert sheet["tab_color"]["rgb"] == "FF3366CC"
    assert sheet["show_gridlines"] is False
    assert sheet["zoom_scale"] == 150
    assert 8 <= sheet["dimensions"]["columns"]["C"]["width"] <= 24

    unmerged_job = temp / "unmerge.json"
    write_json(
        unmerged_job,
        {
            "schema_version": 1,
            "operations": [{"op": "set_sheet", "sheet": "Raw Headers", "unmerge": ["A2:B2"]}],
        },
    )
    unmerged_path = temp / "unmerged.xlsx"
    payload = run_tool("edit", merged_path, unmerged_job, unmerged_path)
    assert payload["result"]["counts"]["ranges_unmerged"] == 1
    inspection = run_tool("inspect", unmerged_path, "--sheet", "Raw Headers", "--no-cells")
    assert inspection["result"]["sheets"][0]["merged_cells"] == []

    bad_unmerge_job = temp / "bad-unmerge.json"
    write_json(
        bad_unmerge_job,
        {
            "schema_version": 1,
            "operations": [{"op": "set_sheet", "sheet": "Raw Headers", "unmerge": ["A9:B9"]}],
        },
    )
    failure = run_tool("edit", source, bad_unmerge_job, temp / "bad-unmerge.xlsx", expect_status=2)
    assert failure["error"]["category"] == "bad_input"
    assert failure["error"]["details"]["range"] == "A9:B9"


def auto_width_create(temp: Path) -> None:
    job_path = temp / "auto-width.json"
    write_json(
        job_path,
        {
            "schema_version": 1,
            "sheets": [
                {
                    "name": "Wide",
                    "rows": [
                        ["Header long enough", "x"],
                        ["Value that is much longer than the header", "y"],
                    ],
                    "auto_width": True,
                }
            ],
        },
    )
    output = temp / "auto-width.xlsx"
    payload = run_tool("create", job_path, output)
    assert payload["result"]["counts"]["columns_auto_sized"] == 2
    inspection = run_tool("inspect", output, "--no-cells")
    columns = inspection["result"]["sheets"][0]["dimensions"]["columns"]
    assert 40 <= columns["A"]["width"] <= 60
    assert columns["B"]["width"] == 10


def formula_error_scan(temp: Path) -> None:
    plain = temp / "plain-errors.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "ok"
    workbook.save(plain)
    workbook.close()
    crafted = temp / "crafted-errors.xlsx"
    with zipfile.ZipFile(plain) as bundle, zipfile.ZipFile(crafted, "w") as target:
        for item in bundle.infolist():
            data = bundle.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                text = data.decode("utf-8")
                assert "</sheetData>" in text
                injected = (
                    '<row r="2">'
                    '<c r="A2" t="e"><v>#REF!</v></c>'
                    '<c r="B2" t="e"><f>1/0</f><v>#DIV/0!</v></c>'
                    "</row></sheetData>"
                )
                data = text.replace("</sheetData>", injected).encode("utf-8")
            target.writestr(item, data)
    inspection = run_tool("inspect", crafted)
    assert inspection["result"]["counts"]["error_cells"] == 2
    sheet = inspection["result"]["sheets"][0]
    assert sheet["errors"]["count"] == 2
    assert sheet["errors"]["by_error"] == {"#DIV/0!": 1, "#REF!": 1}
    kinds = {sample["error"]: sample["kind"] for sample in sheet["errors"]["samples"]}
    assert kinds == {"#REF!": "literal", "#DIV/0!": "cached_formula"}
    assert any("error value(s)" in warning for warning in inspection["warnings"])


def font_inventory(temp: Path, source: Path) -> None:
    inspection = run_tool("inspect", source, "--no-cells")
    fonts = inspection["result"]["workbook"]["fonts"]
    referenced = [name.casefold() for name in fonts["referenced"]]
    assert "calibri" in referenced
    assert fonts["embedded"] == []
    assert fonts["unembedded"] == fonts["referenced"]
    assert any(warning.startswith("RELEASE BLOCKER") for warning in inspection["warnings"])

    arial_job = temp / "arial.json"
    write_json(
        arial_job,
        {
            "schema_version": 1,
            "sheets": [
                {
                    "name": "Portable",
                    "rows": [[{"value": "Arial only", "font": {"name": "Arial"}}]],
                }
            ],
        },
    )
    arial_path = temp / "arial.xlsx"
    run_tool("create", arial_job, arial_path)
    inspection = run_tool("inspect", arial_path, "--no-cells")
    fonts = inspection["result"]["workbook"]["fonts"]
    assert fonts["referenced"] == ["Arial"]
    assert not any(warning.startswith("RELEASE BLOCKER") for warning in inspection["warnings"])


def pdf_render_option_validation(temp: Path, source: Path) -> None:
    failure = run_tool(
        "convert",
        source,
        temp / "option-check.pdf",
        "--sheet",
        "Sales",
        expect_status=2,
    )
    assert failure["error"]["category"] == "bad_input"
    failure = run_tool(
        "convert",
        source,
        temp / "option-check.csv",
        "--timeout",
        "30",
        expect_status=2,
    )
    assert failure["error"]["category"] == "bad_input"
    failure = run_tool(
        "render",
        source,
        temp / "render-bad-dpi",
        "--dpi",
        "9999",
        expect_status=2,
    )
    assert failure["error"]["category"] == "bad_input"


def pdf_and_render(temp: Path, source: Path, require_libreoffice: bool) -> dict[str, Any]:
    libreoffice = shutil.which("soffice") or shutil.which("libreoffice")
    if libreoffice is None:
        if require_libreoffice:
            raise AssertionError("LibreOffice was required but is not available on PATH.")
        return {"status": "skipped", "reason": "LibreOffice is not available on PATH"}
    before = sha256(source)
    pdf = temp / "converted.pdf"
    conversion = run_tool(
        "convert",
        source,
        pdf,
        "--timeout",
        "300",
        timeout_seconds=360,
    )
    pages = conversion["result"]["counts"]["pdf_pages"]
    assert pages >= 1
    assert conversion["result"]["verification"]["pdf_openable"]
    assert conversion["result"]["verification"]["source_unchanged_at_publish_gate"]
    with pdf.open("rb") as handle:
        assert handle.read(5) == b"%PDF-"
    assert sha256(source) == before
    status: dict[str, Any] = {"status": "passed", "executable": libreoffice, "pdf_pages": pages}
    if importlib.util.find_spec("pypdfium2") is None:
        status["render"] = {
            "status": "skipped",
            "reason": "optional pypdfium2 package is not installed",
        }
        return status
    render_dir = temp / "rendered-pages"
    rendered = run_tool(
        "render",
        source,
        render_dir,
        "--timeout",
        "300",
        timeout_seconds=360,
    )
    assert rendered["result"]["counts"]["pages_rendered"] == pages
    files = sorted(render_dir.glob("page-*.png"))
    assert len(files) == pages
    hashes = set()
    for file, record in zip(files, rendered["result"]["images"], strict=True):
        with file.open("rb") as handle:
            assert handle.read(8) == b"\x89PNG\r\n\x1a\n"
        assert record["file"] == file.name
        assert record["width_px"] > 0 and record["height_px"] > 0
        hashes.add(sha256(file))
    if pages > 1:
        assert len(hashes) >= 2
    selected_dir = temp / "rendered-first-page"
    selected = run_tool(
        "render",
        source,
        selected_dir,
        "--pages",
        "1",
        "--timeout",
        "300",
        timeout_seconds=360,
    )
    assert selected["result"]["counts"]["pages_rendered"] == 1
    assert (selected_dir / "page-001.png").is_file()
    failure = run_tool("render", source, render_dir, expect_status=3)
    assert failure["error"]["category"] == "unsupported_operation"
    assert sha256(source) == before
    status["render"] = {"status": "passed", "pages": pages}
    return status


def subprocess_timeout_regression(source: Path) -> None:
    started = time.monotonic()
    try:
        run_tool("inspect", source, timeout_seconds=0.0001)
    except AssertionError as exc:
        assert "CLI subprocess timed out" in str(exc)
    else:
        raise AssertionError("CLI subprocess did not time out under the controlled deadline.")
    assert time.monotonic() - started < 2


def cumulative_json_budget_regression(temp: Path) -> None:
    job_path = temp / "cumulative-budget.json"
    write_json(
        job_path,
        {
            "schema_version": 1,
            "sheets": [
                {"name": "One", "rows": [[1, 2]]},
                {"name": "Two", "rows": [[3, 4]]},
            ],
        },
    )
    failure = run_tool(
        "create",
        job_path,
        temp / "cumulative-budget.xlsx",
        "--cell-limit",
        "3",
        expect_status=6,
    )
    assert failure["error"]["category"] == "resource_limit"
    assert failure["error"]["details"]["cells"] == 4
    assert failure["error"]["details"]["limit"] == 3


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--require-libreoffice",
        action="store_true",
        help="fail unless LibreOffice PDF conversion can be exercised",
    )
    args = parser.parse_args()
    with tempfile.TemporaryDirectory(prefix="xlsx-skill-smoke-") as raw_temp:
        temp = Path(raw_temp)
        created = create_fixture(temp)
        edited = inspect_extract_edit(temp, created)
        summarized = clean_and_summarize(temp, edited)
        round_trip_delimited(temp)
        raw_all_sheet_export(temp, summarized)
        case_insensitive_dependency(temp, created)
        failure_path(temp, created)
        external_link_warning(temp)
        merge_and_sheet_settings(temp, created)
        auto_width_create(temp)
        formula_error_scan(temp)
        font_inventory(temp, created)
        pdf_render_option_validation(temp, created)
        libreoffice_status = pdf_and_render(temp, created, args.require_libreoffice)
        subprocess_timeout_regression(created)
        cumulative_json_budget_regression(temp)
    print(
        json.dumps(
            {
                "schema_version": 1,
                "ok": True,
                "test": "xlsx_smoke",
                "fixtures": "temporary",
                "operations": [
                    "create",
                    "inspect",
                    "extract",
                    "edit",
                    "clean",
                    "summarize",
                    "convert_csv_xlsx_tsv_xlsx",
                    "raw_all_sheet_static_summary_export",
                    "header_aware_export",
                    "conversion_option_direction_validation",
                    "json_formula_preservation",
                    "case_insensitive_formula_dependencies",
                    "bad_input_categories",
                    "xml_encoding_security",
                    "subprocess_timeout",
                    "cumulative_json_cell_budget",
                    "external_link_preservation_warning",
                    "macro_refusal",
                    "set_sheet_visual_settings_merge_unmerge",
                    "auto_width",
                    "formula_error_scan",
                    "font_portability_inventory",
                    "pdf_and_render_option_validation",
                    "xlsx_to_pdf_and_page_render",
                ],
                "libreoffice": libreoffice_status,
            },
            sort_keys=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

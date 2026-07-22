#!/usr/bin/env python3
# SPDX-License-Identifier: Apache-2.0
"""Safely inspect, extract, create, edit, clean, summarize, convert, and render XLSX files."""

from __future__ import annotations

import argparse
import codecs
import csv
import datetime as dt
import hashlib
import json
import math
import os
import re
import shutil
import subprocess
import sys
import tempfile
import traceback
import zipfile
from collections import Counter
from collections.abc import Callable, Iterable, Mapping, Sequence
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Any, NoReturn
from xml.etree import ElementTree

IMPORT_ERROR: ImportError | None = None
try:
    import numpy as np
    import openpyxl
    import pandas as pd
    import pypdf
    from openpyxl import Workbook, load_workbook
    from openpyxl.chart import (
        AreaChart,
        BarChart,
        LineChart,
        PieChart,
        Reference,
        ScatterChart,
        Series,
    )
    from openpyxl.formatting.rule import (
        CellIsRule,
        ColorScaleRule,
        DataBarRule,
        FormulaRule,
        IconSetRule,
    )
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        NamedStyle,
        PatternFill,
        Protection,
        Side,
    )
    from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries
    from openpyxl.utils.cell import quote_sheetname
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from PIL import Image
except ImportError as exc:  # pragma: no cover - exercised only without required dependencies
    IMPORT_ERROR = exc


SCHEMA_VERSION = 1
MAX_PACKAGE_BYTES = 100 * 1024 * 1024
MAX_ZIP_MEMBERS = 10_000
MAX_EXPANDED_BYTES = 512 * 1024 * 1024
MAX_MEMBER_BYTES = 128 * 1024 * 1024
MAX_COMPRESSION_RATIO = 1_000
MAX_INSPECT_CELLS = 100_000
MAX_DATA_CELLS = 2_000_000
XML_DECLARATION = re.compile(r"^\s*<\?xml\b(?P<body>.*?)\?>", re.IGNORECASE | re.DOTALL)
XML_ENCODING = re.compile(
    r"\bencoding\s*=\s*(['\"])(?P<encoding>[^'\"]+)\1",
    re.IGNORECASE,
)
XML_FORBIDDEN = re.compile(r"<!\s*(?:DOCTYPE|ENTITY)\b", re.IGNORECASE)
MACRO_MARKERS = (
    "vbaproject.bin",
    "vbaProjectSignature.bin".lower(),
    "application/vnd.ms-office.vbaproject",
    "application/vnd.ms-excel.sheet.macroenabled",
    "application/vnd.ms-excel.template.macroenabled",
)
DANGEROUS_FORMULA_PREFIXES = ("=", "+", "-", "@")
FORMULA_LEADING_WHITESPACE = " \t\r\n\ufeff\u00a0"
SUPPORTED_WORKBOOK_SUFFIX = ".xlsx"
UNSUPPORTED_WORKBOOK_SUFFIXES = {".xls", ".xlsb", ".xlsm", ".xltm", ".xla", ".xlam"}
FORMULA_ERROR_LITERALS = {
    "#REF!",
    "#DIV/0!",
    "#VALUE!",
    "#NAME?",
    "#N/A",
    "#NULL!",
    "#NUM!",
    "#SPILL!",
    "#CALC!",
}
MAX_FORMULA_ERROR_SAMPLES = 20
PORTABLE_RELEASE_FONTS = {"arial", "times new roman", "courier new"}
MIN_RENDER_DPI = 36
MAX_RENDER_DPI = 300
MAX_RENDER_PIXELS_PER_PAGE = 25_000_000
MAX_RENDER_TOTAL_PIXELS = 250_000_000
MAX_RENDER_PAGES = 200
DEFAULT_EXTERNAL_TOOL_TIMEOUT = 120.0
AUTO_WIDTH_DEFAULTS = {"min_width": 8, "max_width": 60, "sample_rows": 200}
FAILURE_STATUS = {
    "bad_input": 2,
    "unsupported_operation": 3,
    "missing_dependency": 4,
    "ambiguous_edit": 5,
    "resource_limit": 6,
    "licensing_precondition": 7,
    "post_write_validation": 8,
    "internal_error": 9,
    "external_tool_failure": 10,
}


class ToolError(Exception):
    """Expected command failure with a stable category and process status."""

    def __init__(
        self,
        category: str,
        message: str,
        *,
        details: Mapping[str, Any] | None = None,
    ) -> None:
        super().__init__(message)
        self.category = category
        self.details = dict(details or {})

    @property
    def status(self) -> int:
        return FAILURE_STATUS[self.category]


class JsonArgumentParser(argparse.ArgumentParser):
    """Route argparse diagnostics through the JSON error contract."""

    def error(self, message: str) -> NoReturn:
        raise ToolError("bad_input", message)


def fail(category: str, message: str, **details: Any) -> NoReturn:
    raise ToolError(category, message, details=details)


def ensure_dependencies() -> None:
    if IMPORT_ERROR is not None:
        fail(
            "missing_dependency",
            "Install the skill's requirements before running the tool.",
            import_error=str(IMPORT_ERROR),
            requirements="requirements.txt",
        )


def json_default(value: Any) -> Any:
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, np.generic):
        item = value.item()
        if isinstance(item, float) and not math.isfinite(item):
            return None
        return item
    if pd.isna(value):
        return None
    if isinstance(value, set):
        return sorted(value)
    return str(value)


def emit(payload: Mapping[str, Any], *, stream: Any = sys.stdout) -> None:
    print(
        json.dumps(
            payload,
            ensure_ascii=False,
            sort_keys=True,
            default=json_default,
            allow_nan=False,
        ),
        file=stream,
        flush=True,
    )


def versions() -> dict[str, str]:
    return {
        "python": ".".join(map(str, sys.version_info[:3])),
        "openpyxl": openpyxl.__version__,
        "pandas": pd.__version__,
        "numpy": np.__version__,
        "pypdf": pypdf.__version__,
    }


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


def success(operation: str, result: Mapping[str, Any], warnings: Sequence[str]) -> dict[str, Any]:
    return {
        "schema_version": SCHEMA_VERSION,
        "ok": True,
        "operation": operation,
        "result": result,
        "warnings": list(dict.fromkeys(warnings)),
        "versions": versions(),
    }


def load_json_object(path: Path, description: str = "JSON file") -> dict[str, Any]:
    def reject_constant(value: str) -> NoReturn:
        raise ValueError(f"Non-finite JSON constant {value!r} is not permitted.")

    try:
        value = json.loads(
            path.read_text(encoding="utf-8"),
            parse_constant=reject_constant,
        )
    except FileNotFoundError:
        fail("bad_input", f"{description} does not exist.", path=path)
    except UnicodeDecodeError as exc:
        fail("bad_input", f"{description} must be UTF-8.", path=path, reason=str(exc))
    except json.JSONDecodeError as exc:
        fail(
            "bad_input",
            f"{description} is not valid JSON.",
            path=path,
            line=exc.lineno,
            column=exc.colno,
            reason=exc.msg,
        )
    except ValueError as exc:
        fail("bad_input", f"{description} contains an invalid value.", path=path, reason=str(exc))
    except OSError as exc:
        fail("bad_input", f"Could not read {description}.", path=path, reason=str(exc))
    if not isinstance(value, dict):
        fail("bad_input", f"{description} must contain a JSON object.", path=path)
    return value


def validate_job(job: Mapping[str, Any]) -> None:
    version = job.get("schema_version")
    if version != SCHEMA_VERSION:
        fail(
            "bad_input",
            "Unsupported or missing job schema_version.",
            expected=SCHEMA_VERSION,
            actual=version,
        )


def cell_budget_from_args(args: argparse.Namespace) -> CellBudget:
    limit = args.cell_limit
    if not 1 <= limit <= MAX_DATA_CELLS:
        fail(
            "bad_input",
            "--cell-limit must be between 1 and the fixed maximum.",
            minimum=1,
            maximum=MAX_DATA_CELLS,
            actual=limit,
        )
    return CellBudget(limit=limit)


def reject_unknown_keys(
    value: Mapping[str, Any],
    allowed: set[str] | frozenset[str],
    label: str,
) -> None:
    unknown = set(value) - set(allowed)
    if unknown:
        fail(
            "bad_input",
            f"{label} contains unsupported properties.",
            properties=sorted(unknown),
        )


def require_mapping(value: Any, label: str) -> dict[str, Any]:
    if not isinstance(value, dict):
        fail("bad_input", f"{label} must be an object.")
    return value


def require_list(value: Any, label: str) -> list[Any]:
    if not isinstance(value, list):
        fail("bad_input", f"{label} must be an array.")
    return value


def require_string(value: Any, label: str, *, nonempty: bool = True) -> str:
    if not isinstance(value, str) or (nonempty and not value):
        fail("bad_input", f"{label} must be a{' non-empty' if nonempty else ''} string.")
    return value


def check_cell_budget(rows: int, columns: int, *, label: str) -> None:
    count = rows * columns
    if count > MAX_DATA_CELLS:
        fail(
            "resource_limit",
            f"{label} exceeds the cell-processing limit.",
            cells=count,
            limit=MAX_DATA_CELLS,
        )


@dataclass
class CellBudget:
    """Cumulative cell-processing budget for one create or edit job."""

    used: int = 0
    limit: int = MAX_DATA_CELLS

    def consume(self, count: int, *, label: str) -> None:
        if count < 0:
            fail("bad_input", f"{label} produced an invalid negative cell count.")
        attempted = self.used + count
        if attempted > self.limit:
            fail(
                "resource_limit",
                f"{label} exceeds the total cell-processing limit.",
                cells=attempted,
                previous_cells=self.used,
                added_cells=count,
                limit=self.limit,
            )
        self.used = attempted


def workbook_suffix(path: Path) -> str:
    return path.suffix.lower()


def reject_unsupported_workbook_extension(path: Path) -> None:
    suffix = workbook_suffix(path)
    if suffix in UNSUPPORTED_WORKBOOK_SUFFIXES:
        message = (
            "Macro-enabled workbooks are refused."
            if suffix in {".xlsm", ".xltm", ".xla", ".xlam"}
            else "Only the OOXML .xlsx workbook format is supported."
        )
        fail("unsupported_operation", message, path=path, extension=suffix)
    if suffix != SUPPORTED_WORKBOOK_SUFFIX:
        fail(
            "bad_input",
            "Workbook path must use the .xlsx extension.",
            path=path,
            extension=suffix,
        )


def _xml_encoding_from_prefix(content: bytes, member: str) -> str:
    if content.startswith(codecs.BOM_UTF8):
        return "utf-8-sig"
    if content.startswith(codecs.BOM_UTF32_LE):
        return "utf-32-le"
    if content.startswith(codecs.BOM_UTF32_BE):
        return "utf-32-be"
    if content.startswith(codecs.BOM_UTF16_LE):
        return "utf-16-le"
    if content.startswith(codecs.BOM_UTF16_BE):
        return "utf-16-be"
    if content.startswith(b"\x00\x00\x00<"):
        return "utf-32-be"
    if content.startswith(b"<\x00\x00\x00"):
        return "utf-32-le"
    if content.startswith(b"\x00<\x00?"):
        return "utf-16-be"
    if content.startswith(b"<\x00?\x00"):
        return "utf-16-le"
    if content.startswith((b"\x00",)) and content:
        fail("bad_input", "Unsupported or malformed XML encoding.", member=member)
    return "utf-8"


def decode_package_xml(content: bytes, member: str) -> str:
    """Decode one OOXML XML part strictly before checking declarations."""

    encoding = _xml_encoding_from_prefix(content, member)
    try:
        text = content.decode(encoding)
    except (LookupError, UnicodeDecodeError) as exc:
        fail(
            "bad_input",
            "OOXML XML uses an invalid or unsupported encoding.",
            member=member,
            encoding=encoding,
            reason=str(exc),
        )
    text = text.lstrip("\ufeff")
    declaration = XML_DECLARATION.match(text)
    if declaration:
        declared_match = XML_ENCODING.search(declaration.group("body"))
        if declared_match:
            declared = declared_match.group("encoding")
            try:
                canonical = codecs.lookup(declared).name
            except LookupError as exc:
                fail(
                    "bad_input",
                    "OOXML XML declares an unknown encoding.",
                    member=member,
                    encoding=declared,
                    reason=str(exc),
                )
            allowed = {
                "utf-8",
                "utf-8-sig",
                "utf-16",
                "utf-16-le",
                "utf-16-be",
                "utf-32",
                "utf-32-le",
                "utf-32-be",
            }
            if canonical not in allowed:
                fail(
                    "bad_input",
                    "OOXML XML declares an unsupported encoding.",
                    member=member,
                    encoding=declared,
                )
    if XML_FORBIDDEN.search(text):
        fail(
            "unsupported_operation",
            "DTD and entity declarations are refused in OOXML XML.",
            member=member,
        )
    try:
        ElementTree.fromstring(content)
    except (ElementTree.ParseError, ValueError) as exc:
        fail(
            "bad_input",
            "An OOXML XML part is malformed.",
            member=member,
            reason=str(exc),
        )
    return text


def _safe_content_types(content: bytes, path: Path) -> None:
    text = decode_package_xml(content, "[Content_Types].xml")
    lowered = text.casefold()
    if any(marker.casefold() in lowered for marker in MACRO_MARKERS):
        fail("unsupported_operation", "Macro-bearing OOXML packages are refused.", path=path)


def preflight_xlsx(path: Path) -> dict[str, Any]:
    reject_unsupported_workbook_extension(path)
    try:
        stat = path.stat()
    except FileNotFoundError:
        fail("bad_input", "Source workbook does not exist.", path=path)
    except OSError as exc:
        fail("bad_input", "Could not inspect source workbook.", path=path, reason=str(exc))
    if not path.is_file():
        fail("bad_input", "Source workbook must be a regular file.", path=path)
    if stat.st_size > MAX_PACKAGE_BYTES:
        fail(
            "resource_limit",
            "Workbook package exceeds the compressed-size limit.",
            bytes=stat.st_size,
            limit=MAX_PACKAGE_BYTES,
        )
    try:
        with path.open("rb") as handle:
            if handle.read(4) != b"PK\x03\x04":
                fail("bad_input", "Workbook does not have an OOXML ZIP signature.", path=path)
        with zipfile.ZipFile(path) as package:
            members = package.infolist()
            if len(members) > MAX_ZIP_MEMBERS:
                fail(
                    "resource_limit",
                    "Workbook has too many ZIP members.",
                    members=len(members),
                    limit=MAX_ZIP_MEMBERS,
                )
            names = {info.filename for info in members}
            required = {"[Content_Types].xml", "xl/workbook.xml"}
            if not required.issubset(names):
                fail(
                    "bad_input",
                    "ZIP package is missing required XLSX members.",
                    missing=sorted(required - names),
                )
            expanded = 0
            for info in members:
                normalized = Path(info.filename.replace("\\", "/"))
                if (
                    info.filename.startswith("/")
                    or ".." in normalized.parts
                    or "\x00" in info.filename
                ):
                    fail("bad_input", "Unsafe path in workbook ZIP member.", member=info.filename)
                if info.flag_bits & 0x1:
                    fail(
                        "unsupported_operation",
                        "Encrypted ZIP members are not supported.",
                        member=info.filename,
                    )
                if info.file_size > MAX_MEMBER_BYTES:
                    fail(
                        "resource_limit",
                        "Workbook ZIP member exceeds the expanded-size limit.",
                        member=info.filename,
                        bytes=info.file_size,
                        limit=MAX_MEMBER_BYTES,
                    )
                expanded += info.file_size
                if expanded > MAX_EXPANDED_BYTES:
                    fail(
                        "resource_limit",
                        "Workbook exceeds the total expanded-size limit.",
                        bytes=expanded,
                        limit=MAX_EXPANDED_BYTES,
                    )
                ratio = info.file_size / max(info.compress_size, 1)
                if info.file_size > 1_000_000 and ratio > MAX_COMPRESSION_RATIO:
                    fail(
                        "resource_limit",
                        "Workbook ZIP member has an unsafe compression ratio.",
                        member=info.filename,
                        ratio=round(ratio, 1),
                        limit=MAX_COMPRESSION_RATIO,
                    )
                lower_name = info.filename.lower()
                if any(marker in lower_name for marker in MACRO_MARKERS[:2]):
                    fail(
                        "unsupported_operation",
                        "Macro-bearing OOXML packages are refused.",
                        member=info.filename,
                    )
                if lower_name.endswith((".xml", ".rels")):
                    content = package.read(info)
                    decode_package_xml(content, info.filename)
            _safe_content_types(package.read("[Content_Types].xml"), path)
    except zipfile.BadZipFile as exc:
        fail("bad_input", "Workbook is not a valid ZIP package.", path=path, reason=str(exc))
    except (NotImplementedError, RuntimeError) as exc:
        fail(
            "bad_input",
            "Workbook ZIP members could not be decoded.",
            path=path,
            reason=str(exc),
        )
    except OSError as exc:
        fail("bad_input", "Could not read workbook package.", path=path, reason=str(exc))
    return {
        "compressed_bytes": stat.st_size,
        "expanded_bytes": expanded,
        "zip_members": len(members),
        "macros": False,
    }


def open_workbook(
    path: Path,
    *,
    data_only: bool = False,
    read_only: bool = False,
) -> Any:
    preflight_xlsx(path)
    try:
        return load_workbook(
            path,
            data_only=data_only,
            read_only=read_only,
            keep_links=True,
            keep_vba=False,
        )
    except (OSError, ValueError, KeyError, ElementTree.ParseError) as exc:
        fail("bad_input", "openpyxl could not open the workbook.", path=path, reason=str(exc))


def require_output_path(
    output: Path,
    *,
    source: Path | None,
    overwrite: bool,
    suffix: str | None = None,
) -> None:
    if suffix is not None and output.suffix.lower() != suffix:
        fail(
            "bad_input",
            f"Output path must use the {suffix} extension.",
            path=output,
        )
    try:
        same_as_source = source is not None and output.resolve() == source.resolve()
    except OSError:
        same_as_source = False
    if same_as_source and not overwrite:
        fail(
            "bad_input",
            "Source and destination must be distinct unless --overwrite is explicit.",
            path=output,
        )
    if output.exists() and not overwrite:
        fail(
            "bad_input",
            "Destination already exists; pass --overwrite to replace it.",
            path=output,
        )
    if not output.parent.exists() or not output.parent.is_dir():
        fail("bad_input", "Destination parent directory does not exist.", path=output.parent)


def temporary_sibling(path: Path, suffix: str | None = None) -> Path:
    descriptor, raw = tempfile.mkstemp(
        prefix=f".{path.name}.",
        suffix=suffix or path.suffix,
        dir=path.parent,
    )
    os.close(descriptor)
    temp = Path(raw)
    temp.unlink()
    return temp


def atomic_publish_file(temp: Path, output: Path) -> None:
    try:
        with temp.open("rb") as handle:
            os.fsync(handle.fileno())
        os.replace(temp, output)
        try:
            directory_fd = os.open(output.parent, os.O_RDONLY)
        except OSError:
            return
        try:
            os.fsync(directory_fd)
        finally:
            os.close(directory_fd)
    except OSError as exc:
        temp.unlink(missing_ok=True)
        fail(
            "post_write_validation",
            "Could not atomically publish the verified output.",
            path=output,
            reason=str(exc),
        )


def verify_workbook(path: Path, expected_sheets: Sequence[str] | None = None) -> dict[str, Any]:
    try:
        package = preflight_xlsx(path)
    except ToolError as exc:
        fail(
            "post_write_validation",
            "Saved workbook failed OOXML package verification.",
            original_category=exc.category,
            original_message=str(exc),
            **exc.details,
        )
    try:
        workbook = load_workbook(path, read_only=False, data_only=False, keep_links=True)
        sheet_names = workbook.sheetnames
        if expected_sheets is not None and sheet_names != list(expected_sheets):
            fail(
                "post_write_validation",
                "Reopened workbook sheet order differs from the saved workbook.",
                expected=list(expected_sheets),
                actual=sheet_names,
            )
        workbook.close()
    except ToolError:
        raise
    except Exception as exc:
        fail(
            "post_write_validation",
            "Saved workbook could not be reopened.",
            path=path,
            reason=str(exc),
        )
    return {"reopened": True, "sheet_names": sheet_names, "package": package}


def save_workbook_atomic(
    workbook: Any,
    output: Path,
    *,
    source: Path | None,
    overwrite: bool,
) -> dict[str, Any]:
    require_output_path(output, source=source, overwrite=overwrite, suffix=".xlsx")
    temp = temporary_sibling(output, ".xlsx")
    expected_sheets = list(workbook.sheetnames)
    try:
        workbook.save(temp)
        verification = verify_workbook(temp, expected_sheets)
        atomic_publish_file(temp, output)
        return verification
    except ToolError:
        temp.unlink(missing_ok=True)
        raise
    except Exception as exc:
        temp.unlink(missing_ok=True)
        fail(
            "post_write_validation",
            "Could not save workbook.",
            path=output,
            reason=str(exc),
        )


def write_text_atomic(
    output: Path,
    writer: Callable[[Any], None],
    *,
    source: Path | None,
    overwrite: bool,
    encoding: str,
) -> dict[str, Any]:
    try:
        codecs.lookup(encoding)
    except LookupError as exc:
        fail("bad_input", "Unknown text encoding.", encoding=encoding, reason=str(exc))
    require_output_path(output, source=source, overwrite=overwrite)
    temp = temporary_sibling(output)
    try:
        with temp.open("w", encoding=encoding, newline="") as handle:
            writer(handle)
            handle.flush()
            os.fsync(handle.fileno())
        with temp.open("r", encoding=encoding, newline="") as handle:
            handle.read(1)
        size = temp.stat().st_size
        atomic_publish_file(temp, output)
        return {"reopened": True, "bytes": size, "encoding": encoding}
    except ToolError:
        temp.unlink(missing_ok=True)
        raise
    except (OSError, UnicodeError, csv.Error) as exc:
        temp.unlink(missing_ok=True)
        fail(
            "post_write_validation",
            "Could not write or verify text output.",
            path=output,
            reason=str(exc),
        )


def worksheet_by_name(workbook: Any, name: str | None) -> Any:
    if name is None:
        return workbook.active
    if name not in workbook.sheetnames:
        fail("bad_input", "Worksheet does not exist.", sheet=name)
    return workbook[name]


def normalize_range(worksheet: Any, range_text: str | None) -> tuple[int, int, int, int]:
    if range_text is None:
        min_row = worksheet.min_row
        min_col = worksheet.min_column
        max_row = worksheet.max_row
        max_col = worksheet.max_column
    else:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(range_text)
        except (TypeError, ValueError) as exc:
            fail("bad_input", "Invalid A1 cell range.", range=range_text, reason=str(exc))
    if not all(isinstance(value, int) for value in (min_row, min_col, max_row, max_col)):
        fail(
            "bad_input",
            "Range must include bounded rows and columns, such as A1:D20.",
            range=range_text,
        )
    if min_row < 1 or min_col < 1 or max_row < min_row or max_col < min_col:
        fail("bad_input", "Invalid worksheet range bounds.", range=range_text)
    check_cell_budget(max_row - min_row + 1, max_col - min_col + 1, label="Worksheet range")
    return min_row, min_col, max_row, max_col


def range_to_a1(bounds: tuple[int, int, int, int]) -> str:
    min_row, min_col, max_row, max_col = bounds
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def iter_range_values(
    worksheet: Any,
    bounds: tuple[int, int, int, int],
) -> Iterable[list[Any]]:
    min_row, min_col, max_row, max_col = bounds
    for row in worksheet.iter_rows(
        min_row=min_row,
        min_col=min_col,
        max_row=max_row,
        max_col=max_col,
    ):
        yield [cell.value for cell in row]


def literal_cell_value(cell: Any, value: Any, *, allow_formula: bool = False) -> None:
    if isinstance(value, dict):
        fail(
            "bad_input",
            "Cell scalar values cannot be JSON objects; use a cell specification.",
            coordinate=cell.coordinate,
        )
    cell.value = value
    if isinstance(value, str) and value.startswith("=") and not allow_formula:
        cell.data_type = "s"


def formula_cell_value(cell: Any, formula: str) -> None:
    require_string(formula, "formula")
    if not formula.startswith("="):
        fail("bad_input", "Formula values must begin with '='.", coordinate=cell.coordinate)
    cell.value = formula


def color_value(value: Any, label: str) -> str | None:
    if value is None:
        return None
    text = require_string(value, label)
    if not re.fullmatch(r"(?:[0-9A-Fa-f]{6}|[0-9A-Fa-f]{8})", text):
        fail("bad_input", f"{label} must be a 6- or 8-digit hexadecimal color.")
    text = text.upper()
    # openpyxl pads 6-digit values with a fully transparent 00 alpha; make them opaque.
    if len(text) == 6:
        text = "FF" + text
    return text


def build_font(spec: Mapping[str, Any] | None) -> Any:
    data = dict(spec or {})
    if "color" in data:
        data["color"] = color_value(data["color"], "font.color")
    allowed = {
        "name",
        "size",
        "bold",
        "italic",
        "vertAlign",
        "underline",
        "strike",
        "color",
        "outline",
        "shadow",
        "condense",
        "extend",
    }
    unknown = set(data) - allowed
    if unknown:
        fail("bad_input", "Unsupported font property.", properties=sorted(unknown))
    if "size" in data:
        data["sz"] = data.pop("size")
    return Font(**data)


def build_fill(spec: Mapping[str, Any] | None) -> Any:
    data = dict(spec or {})
    fill_type = data.pop("fill_type", data.pop("type", "solid" if data else None))
    fg = color_value(data.pop("fg_color", data.pop("color", None)), "fill.fg_color")
    bg = color_value(data.pop("bg_color", None), "fill.bg_color")
    if data:
        fail("bad_input", "Unsupported fill property.", properties=sorted(data))
    arguments = {"fill_type": fill_type}
    if fg is not None:
        arguments["fgColor"] = fg
    if bg is not None:
        arguments["bgColor"] = bg
    return PatternFill(**arguments)


def build_side(spec: Mapping[str, Any] | None) -> Any:
    data = dict(spec or {})
    style = data.pop("style", None)
    color = color_value(data.pop("color", None), "border color")
    if data:
        fail("bad_input", "Unsupported border-side property.", properties=sorted(data))
    return Side(style=style, color=color)


def build_border(spec: Mapping[str, Any] | None) -> Any:
    data = dict(spec or {})
    allowed = {"left", "right", "top", "bottom", "diagonal", "vertical", "horizontal"}
    unknown = set(data) - allowed
    if unknown:
        fail("bad_input", "Unsupported border property.", properties=sorted(unknown))
    return Border(**{name: build_side(value) for name, value in data.items()})


def build_alignment(spec: Mapping[str, Any] | None) -> Any:
    data = dict(spec or {})
    aliases = {
        "wrap_text": "wrapText",
        "shrink_to_fit": "shrinkToFit",
        "text_rotation": "textRotation",
    }
    for old, new in aliases.items():
        if old in data:
            data[new] = data.pop(old)
    allowed = {
        "horizontal",
        "vertical",
        "textRotation",
        "wrapText",
        "shrinkToFit",
        "indent",
        "relativeIndent",
        "justifyLastLine",
        "readingOrder",
    }
    unknown = set(data) - allowed
    if unknown:
        fail("bad_input", "Unsupported alignment property.", properties=sorted(unknown))
    return Alignment(**data)


def build_protection(spec: Mapping[str, Any] | None) -> Any:
    data = dict(spec or {})
    unknown = set(data) - {"locked", "hidden"}
    if unknown:
        fail("bad_input", "Unsupported protection property.", properties=sorted(unknown))
    return Protection(**data)


def apply_style(cell: Any, spec: Mapping[str, Any]) -> None:
    data = dict(spec)
    if "named_style" in data or "style" in data:
        cell.style = data.pop("named_style", data.pop("style", None))
    if "font" in data:
        cell.font = build_font(require_mapping(data.pop("font"), "font"))
    if "fill" in data:
        cell.fill = build_fill(require_mapping(data.pop("fill"), "fill"))
    if "border" in data:
        cell.border = build_border(require_mapping(data.pop("border"), "border"))
    if "alignment" in data:
        cell.alignment = build_alignment(require_mapping(data.pop("alignment"), "alignment"))
    if "protection" in data:
        cell.protection = build_protection(require_mapping(data.pop("protection"), "protection"))
    if "number_format" in data:
        cell.number_format = require_string(data.pop("number_format"), "number_format")
    if data:
        fail("bad_input", "Unsupported cell-format property.", properties=sorted(data))


def copy_cell_style(source: Any, target: Any) -> None:
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.number_format = source.number_format
    target.protection = copy(source.protection)


def set_cell_from_spec(cell: Any, raw_spec: Any) -> None:
    if not isinstance(raw_spec, dict):
        literal_cell_value(cell, raw_spec)
        return
    spec = dict(raw_spec)
    has_value = "value" in spec
    has_formula = "formula" in spec
    if has_value and has_formula:
        fail(
            "bad_input",
            "Cell specification cannot contain both value and formula.",
            coordinate=cell.coordinate,
        )
    if has_formula:
        formula_cell_value(cell, spec.pop("formula"))
    elif has_value:
        literal_cell_value(
            cell,
            spec.pop("value"),
            allow_formula=bool(spec.pop("allow_formula", False)),
        )
    if "comment" in spec:
        fail("unsupported_operation", "Creating comments is not supported in schema version 1.")
    style_keys = {
        "style",
        "named_style",
        "font",
        "fill",
        "border",
        "alignment",
        "protection",
        "number_format",
    }
    style_spec = {key: spec.pop(key) for key in list(spec) if key in style_keys}
    if style_spec:
        apply_style(cell, style_spec)
    if "hyperlink" in spec:
        hyperlink = require_string(spec.pop("hyperlink"), "hyperlink")
        if not re.match(r"^(?:https?|mailto):", hyperlink, re.IGNORECASE):
            fail("bad_input", "Only http, https, and mailto hyperlinks are accepted.")
        cell.hyperlink = hyperlink
    if spec:
        fail(
            "bad_input",
            "Unsupported cell specification property.",
            coordinate=cell.coordinate,
            properties=sorted(spec),
        )


def serialize_cell_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return None
        return value
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    return str(value)


def serialize_color(color: Any) -> dict[str, Any] | None:
    if color is None:
        return None
    return {
        "type": color.type,
        "rgb": color.rgb if color.type == "rgb" else None,
        "indexed": color.indexed if color.type == "indexed" else None,
        "theme": color.theme if color.type == "theme" else None,
        "tint": color.tint,
    }


def chart_title(chart: Any) -> str | None:
    title = getattr(chart, "title", None)
    if title is None:
        return None
    try:
        paragraphs = title.tx.rich.p
        parts = []
        for paragraph in paragraphs:
            for run in paragraph.r:
                if run.t:
                    parts.append(run.t)
            if paragraph.endParaRPr and getattr(paragraph.endParaRPr, "t", None):
                parts.append(paragraph.endParaRPr.t)
        return "".join(parts) or None
    except (AttributeError, TypeError):
        return str(title)


def chart_anchor(chart: Any) -> str | None:
    anchor = getattr(chart, "anchor", None)
    marker = getattr(anchor, "_from", None)
    if marker is None:
        return anchor if isinstance(anchor, str) else None
    return f"{get_column_letter(marker.col + 1)}{marker.row + 1}"


def inventory_defined_names(workbook: Any) -> list[dict[str, Any]]:
    names: list[dict[str, Any]] = []
    for name, defined in workbook.defined_names.items():
        names.append(
            {
                "name": name,
                "text": getattr(defined, "attr_text", None),
                "local_sheet_id": getattr(defined, "localSheetId", None),
                "hidden": getattr(defined, "hidden", None),
                "type": getattr(defined, "type", None),
            }
        )
    return names


def inventory_external_links(workbook: Any) -> list[dict[str, Any]]:
    links: list[dict[str, Any]] = []
    for index, link in enumerate(getattr(workbook, "_external_links", [])):
        file_link = getattr(link, "file_link", None)
        links.append(
            {
                "index": index,
                "relationship_id": getattr(file_link, "Id", None),
                "target": getattr(file_link, "Target", None),
            }
        )
    return links


def mutation_preservation_warnings(workbook: Any, operation: str) -> list[str]:
    warnings = [
        f"{operation} rewrites the OOXML package; unsupported extension content may be "
        "altered or dropped and must be independently verified."
    ]
    links = inventory_external_links(workbook)
    if links:
        warnings.append(
            f"{operation} is saving a workbook with {len(links)} external link(s); link "
            "targets, cached values, and preservation are not repaired or guaranteed."
        )
    if any(getattr(worksheet, "_pivots", []) for worksheet in workbook.worksheets):
        warnings.append(
            "Existing pivots were not edited; pivot caches, refresh behavior, and "
            "preservation are not guaranteed."
        )
    return warnings


def inventory_conditional_formats(worksheet: Any) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for conditional, rules in worksheet.conditional_formatting._cf_rules.items():
        result.append(
            {
                "range": str(conditional.sqref),
                "rules": [
                    {
                        "type": rule.type,
                        "operator": getattr(rule, "operator", None),
                        "formula": list(getattr(rule, "formula", []) or []),
                        "priority": getattr(rule, "priority", None),
                        "stop_if_true": getattr(rule, "stopIfTrue", None),
                    }
                    for rule in rules
                ],
            }
        )
    return result


def inventory_validations(worksheet: Any) -> list[dict[str, Any]]:
    container = getattr(worksheet, "data_validations", None)
    validations = getattr(container, "dataValidation", []) if container else []
    return [
        {
            "range": str(validation.sqref),
            "type": validation.type,
            "operator": validation.operator,
            "formula1": validation.formula1,
            "formula2": validation.formula2,
            "allow_blank": validation.allow_blank,
            "error": validation.error,
            "prompt": validation.prompt,
        }
        for validation in validations
    ]


def inventory_tables(worksheet: Any) -> list[dict[str, Any]]:
    tables: list[dict[str, Any]] = []
    for table in worksheet.tables.values():
        style = table.tableStyleInfo
        tables.append(
            {
                "name": table.displayName,
                "range": table.ref,
                "totals_row_shown": table.totalsRowShown,
                "style": (
                    {
                        "name": style.name,
                        "show_first_column": style.showFirstColumn,
                        "show_last_column": style.showLastColumn,
                        "show_row_stripes": style.showRowStripes,
                        "show_column_stripes": style.showColumnStripes,
                    }
                    if style
                    else None
                ),
            }
        )
    return tables


def inventory_charts(worksheet: Any) -> list[dict[str, Any]]:
    return [
        {
            "index": index,
            "type": type(chart).__name__,
            "title": chart_title(chart),
            "anchor": chart_anchor(chart),
            "series_count": len(getattr(chart, "series", [])),
            "width": getattr(chart, "width", None),
            "height": getattr(chart, "height", None),
        }
        for index, chart in enumerate(getattr(worksheet, "_charts", []))
    ]


def inventory_pivots(worksheet: Any) -> list[dict[str, Any]]:
    pivots: list[dict[str, Any]] = []
    for index, pivot in enumerate(getattr(worksheet, "_pivots", [])):
        cache = getattr(pivot, "cacheDefinition", None)
        source = getattr(cache, "cacheSource", None)
        pivots.append(
            {
                "index": index,
                "name": getattr(pivot, "name", None),
                "cache_id": getattr(pivot, "cacheId", None),
                "source_type": getattr(source, "type", None),
            }
        )
    return pivots


def inspect_page_setup(worksheet: Any) -> dict[str, Any]:
    setup = worksheet.page_setup
    properties = getattr(worksheet.sheet_properties, "pageSetUpPr", None)
    try:
        print_area = worksheet.print_area
    except (TypeError, ValueError):
        print_area = None
    return {
        "orientation": getattr(setup, "orientation", None),
        "paper_size": getattr(setup, "paperSize", None),
        "scale": getattr(setup, "scale", None),
        "fit_to_width": getattr(setup, "fitToWidth", None),
        "fit_to_height": getattr(setup, "fitToHeight", None),
        "fit_to_page": getattr(properties, "fitToPage", None) if properties else None,
        "print_area": print_area or None,
        "print_title_rows": getattr(worksheet, "print_title_rows", None),
        "print_title_cols": getattr(worksheet, "print_title_cols", None),
    }


def _package_font_scan(path: Path) -> tuple[set[str], dict[str, str | None]]:
    """Collect chart and theme typefaces from already-preflighted package XML."""

    chart_fonts: set[str] = set()
    theme: dict[str, str | None] = {"major": None, "minor": None}
    try:
        with zipfile.ZipFile(path) as package:
            for info in package.infolist():
                lowered = info.filename.lower()
                if not lowered.endswith(".xml"):
                    continue
                is_chart = lowered.startswith("xl/charts/")
                is_theme = lowered.startswith("xl/theme/")
                if not (is_chart or is_theme):
                    continue
                text = decode_package_xml(package.read(info), info.filename)
                root = ElementTree.fromstring(text)
                if is_chart:
                    for element in root.iter():
                        if element.tag.rsplit("}", 1)[-1] in {"latin", "ea", "cs"}:
                            typeface = element.get("typeface")
                            if typeface and not typeface.startswith("+"):
                                chart_fonts.add(typeface)
                    continue
                for scheme_key, tag in (("major", "majorFont"), ("minor", "minorFont")):
                    for element in root.iter():
                        if element.tag.rsplit("}", 1)[-1] != tag:
                            continue
                        for child in element:
                            if child.tag.rsplit("}", 1)[-1] == "latin":
                                typeface = child.get("typeface")
                                if typeface:
                                    theme[scheme_key] = typeface
                        break
    except (OSError, zipfile.BadZipFile, ElementTree.ParseError) as exc:
        fail("bad_input", "Could not scan package fonts.", path=path, reason=str(exc))
    return chart_fonts, theme


def workbook_font_inventory(workbook: Any, path: Path) -> dict[str, Any]:
    """Fonts effectively referenced by non-empty cells and chart text.

    XLSX packages cannot embed fonts, so every referenced font is also unembedded and
    must be installed on each intended renderer. Named-style and theme fonts are
    reported as separate sources only: a named style applied to any cell already shows
    up through that cell's resolved font, and an unapplied one renders nothing. Fonts
    referenced only by unsupported parts (for example rich-text runs in shared strings)
    are not detected.
    """

    cell_fonts: dict[str, str] = {}
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                name = getattr(cell.font, "name", None)
                if name:
                    cell_fonts.setdefault(name.casefold(), name)
    named_style_fonts: dict[str, str] = {}
    for style in getattr(workbook, "_named_styles", []):
        name = getattr(getattr(style, "font", None), "name", None)
        if name:
            named_style_fonts.setdefault(name.casefold(), name)
    scanned_chart_fonts, theme = _package_font_scan(path)
    chart_fonts: dict[str, str] = {}
    for name in scanned_chart_fonts:
        chart_fonts.setdefault(name.casefold(), name)
    referenced: dict[str, str] = {}
    for source in (cell_fonts, chart_fonts):
        for key, value in source.items():
            referenced.setdefault(key, value)
    ordered = sorted(referenced.values(), key=str.casefold)
    return {
        "referenced": ordered,
        "by_source": {
            "cells": sorted(cell_fonts.values(), key=str.casefold),
            "named_styles": sorted(named_style_fonts.values(), key=str.casefold),
            "charts": sorted(chart_fonts.values(), key=str.casefold),
        },
        "theme": theme,
        "embedded": [],
        "unembedded": ordered,
    }


def font_portability_warnings(fonts: Mapping[str, Any]) -> list[str]:
    warnings: list[str] = []
    referenced = list(fonts.get("referenced", []))
    if not referenced:
        return warnings
    warnings.append(
        "Workbook references fonts that renderers must supply ("
        + ", ".join(referenced)
        + "). XLSX cannot embed fonts; a renderer without them substitutes metrics and can "
        "change column fit, wrapping, and print pagination. Do not claim visual fidelity "
        "from structural checks alone."
    )
    nonportable = [name for name in referenced if name.casefold() not in PORTABLE_RELEASE_FONTS]
    if nonportable:
        warnings.append(
            "RELEASE BLOCKER: workbook references fonts outside the conservative "
            "cross-renderer set (Arial, Times New Roman, Courier New): "
            + ", ".join(nonportable)
            + ". Replace these fonts before releasing matching XLSX and PDF deliverables; "
            "renderer substitution can change layout and pagination."
        )
    return warnings


def inspect_sheet(
    worksheet: Any,
    cached_worksheet: Any,
    *,
    include_cells: bool,
    max_cells: int,
    selected_range: str | None,
) -> tuple[dict[str, Any], list[str]]:
    warnings: list[str] = []
    bounds = normalize_range(worksheet, selected_range)
    min_row, min_col, max_row, max_col = bounds
    cell_span = (max_row - min_row + 1) * (max_col - min_col + 1)
    cells: list[dict[str, Any]] = []
    formula_count = 0
    styled_count = 0
    nonempty_count = 0
    style_counts: Counter[tuple[int, str]] = Counter()
    style_samples: dict[tuple[int, str], dict[str, Any]] = {}
    error_counts: Counter[str] = Counter()
    error_samples: list[dict[str, Any]] = []
    truncated = False
    for row in worksheet.iter_rows(
        min_row=min_row,
        min_col=min_col,
        max_row=max_row,
        max_col=max_col,
    ):
        for cell in row:
            if cell.value is None and not cell.has_style:
                continue
            nonempty_count += cell.value is not None
            styled_count += cell.has_style
            if cell.has_style:
                style_key = (cell.style_id, cell.number_format)
                style_counts[style_key] += 1
                style_samples.setdefault(
                    style_key,
                    {
                        "font": {
                            "name": cell.font.name,
                            "size": cell.font.sz,
                            "bold": cell.font.bold,
                            "italic": cell.font.italic,
                            "underline": cell.font.underline,
                            "color": serialize_color(cell.font.color),
                        },
                        "fill": {
                            "type": cell.fill.fill_type,
                            "foreground": serialize_color(cell.fill.fgColor),
                            "background": serialize_color(cell.fill.bgColor),
                        },
                        "border": {
                            "left": getattr(cell.border.left, "style", None),
                            "right": getattr(cell.border.right, "style", None),
                            "top": getattr(cell.border.top, "style", None),
                            "bottom": getattr(cell.border.bottom, "style", None),
                        },
                        "alignment": {
                            "horizontal": cell.alignment.horizontal,
                            "vertical": cell.alignment.vertical,
                            "wrap_text": cell.alignment.wrap_text,
                            "text_rotation": cell.alignment.text_rotation,
                        },
                        "protection": {
                            "locked": cell.protection.locked,
                            "hidden": cell.protection.hidden,
                        },
                    },
                )
            is_formula = cell.data_type == "f"
            formula_count += is_formula
            cached_value: Any = None
            if is_formula:
                cached_value = cached_worksheet[cell.coordinate].value
            error_text: str | None = None
            error_kind: str | None = None
            if cell.data_type == "e" and cell.value is not None:
                error_text, error_kind = str(cell.value), "literal"
            elif (
                is_formula
                and isinstance(cached_value, str)
                and cached_value in FORMULA_ERROR_LITERALS
            ):
                error_text, error_kind = cached_value, "cached_formula"
            if error_text is not None:
                error_counts[error_text] += 1
                if len(error_samples) < MAX_FORMULA_ERROR_SAMPLES:
                    error_samples.append(
                        {
                            "coordinate": cell.coordinate,
                            "error": error_text,
                            "kind": error_kind,
                        }
                    )
            if include_cells and len(cells) < max_cells:
                record = {
                    "coordinate": cell.coordinate,
                    "value": serialize_cell_value(cell.value),
                    "data_type": cell.data_type,
                    "style_id": cell.style_id,
                    "number_format": cell.number_format,
                }
                if is_formula:
                    record["formula"] = cell.value
                    record["cached_value"] = serialize_cell_value(cached_value)
                cells.append(record)
            elif include_cells:
                truncated = True
    if truncated:
        warnings.append(
            f"Cell inventory for sheet {worksheet.title!r} was truncated at {max_cells} records."
        )
    if formula_count:
        warnings.append(
            f"Sheet {worksheet.title!r} contains formulas; cached results may be stale or absent."
        )
    error_count = sum(error_counts.values())
    if error_count:
        breakdown = ", ".join(f"{name} x{count}" for name, count in sorted(error_counts.items()))
        warnings.append(
            f"Sheet {worksheet.title!r} contains {error_count} error value(s): {breakdown}."
        )
    dimensions = {
        "columns": {
            key: {
                "width": value.width,
                "hidden": value.hidden,
                "outline_level": value.outlineLevel,
            }
            for key, value in worksheet.column_dimensions.items()
            if value.width is not None or value.hidden or value.outlineLevel
        },
        "rows": {
            str(key): {
                "height": value.height,
                "hidden": value.hidden,
                "outline_level": value.outlineLevel,
            }
            for key, value in worksheet.row_dimensions.items()
            if value.height is not None or value.hidden or value.outlineLevel
        },
    }
    return (
        {
            "name": worksheet.title,
            "visibility": worksheet.sheet_state,
            "used_range": worksheet.calculate_dimension(),
            "selected_range": range_to_a1(bounds),
            "range_cell_span": cell_span,
            "nonempty_cells": nonempty_count,
            "styled_cells": styled_count,
            "formula_cells": formula_count,
            "freeze_panes": (
                worksheet.freeze_panes.coordinate
                if hasattr(worksheet.freeze_panes, "coordinate")
                else worksheet.freeze_panes
            ),
            "merged_cells": [str(item) for item in worksheet.merged_cells.ranges],
            "auto_filter": worksheet.auto_filter.ref,
            "tab_color": serialize_color(worksheet.sheet_properties.tabColor),
            "show_gridlines": worksheet.sheet_view.showGridLines,
            "zoom_scale": worksheet.sheet_view.zoomScale,
            "page_setup": inspect_page_setup(worksheet),
            "errors": {
                "count": error_count,
                "by_error": dict(sorted(error_counts.items())),
                "samples": error_samples,
            },
            "tables": inventory_tables(worksheet),
            "charts": inventory_charts(worksheet),
            "pivots": inventory_pivots(worksheet),
            "conditional_formatting": inventory_conditional_formats(worksheet),
            "data_validations": inventory_validations(worksheet),
            "dimensions": dimensions,
            "style_inventory": [
                {
                    "style_id": style_id,
                    "number_format": number_format,
                    "cell_count": count,
                    **style_samples[(style_id, number_format)],
                }
                for (style_id, number_format), count in sorted(style_counts.items())
            ],
            "cells": cells if include_cells else None,
            "cell_inventory_truncated": truncated,
        },
        warnings,
    )


def handle_inspect(args: argparse.Namespace) -> dict[str, Any]:
    package = preflight_xlsx(args.source)
    workbook = open_workbook(args.source, data_only=False)
    cached = open_workbook(args.source, data_only=True)
    selected_names = [args.sheet] if args.sheet else workbook.sheetnames
    if args.sheet and args.sheet not in workbook.sheetnames:
        fail("bad_input", "Worksheet does not exist.", sheet=args.sheet)
    max_cells = args.max_cells
    if max_cells < 0 or max_cells > MAX_INSPECT_CELLS:
        fail(
            "resource_limit",
            "--max-cells is outside the supported range.",
            maximum=MAX_INSPECT_CELLS,
        )
    sheets: list[dict[str, Any]] = []
    warnings: list[str] = []
    for name in selected_names:
        inventory, sheet_warnings = inspect_sheet(
            workbook[name],
            cached[name],
            include_cells=not args.no_cells,
            max_cells=max_cells,
            selected_range=args.range,
        )
        sheets.append(inventory)
        warnings.extend(sheet_warnings)
    external_links = inventory_external_links(workbook)
    if external_links:
        warnings.append(
            "External links are inventoried but not repaired or proven to remain resolvable."
        )
    fonts = workbook_font_inventory(workbook, args.source)
    warnings.extend(font_portability_warnings(fonts))
    pivot_count = sum(len(sheet["pivots"]) for sheet in sheets)
    if pivot_count:
        warnings.append(
            "Native pivot tables are inventory-only; edits and refresh behavior are not supported."
        )
    result = {
        "source": str(args.source.resolve()),
        "package": package,
        "workbook": {
            "sheet_order": workbook.sheetnames,
            "active_sheet": workbook.active.title,
            "defined_names": inventory_defined_names(workbook),
            "external_links": external_links,
            "fonts": fonts,
            "macros": False,
            "calculation": {
                "mode": workbook.calculation.calcMode,
                "full_calc_on_load": workbook.calculation.fullCalcOnLoad,
                "force_full_calc": workbook.calculation.forceFullCalc,
            },
            "properties": {
                "title": workbook.properties.title,
                "subject": workbook.properties.subject,
                "creator": workbook.properties.creator,
                "keywords": workbook.properties.keywords,
                "description": workbook.properties.description,
                "category": workbook.properties.category,
                "created": workbook.properties.created,
                "modified": workbook.properties.modified,
            },
        },
        "sheets": sheets,
        "counts": {
            "sheets": len(workbook.sheetnames),
            "tables": sum(len(sheet["tables"]) for sheet in sheets),
            "charts": sum(len(sheet["charts"]) for sheet in sheets),
            "pivots": pivot_count,
            "formulas": sum(sheet["formula_cells"] for sheet in sheets),
            "error_cells": sum(sheet["errors"]["count"] for sheet in sheets),
        },
    }
    workbook.close()
    cached.close()
    return success("inspect", result, warnings)


def sanitize_formula_text(value: Any, *, allow_formulas: bool) -> tuple[Any, bool]:
    if allow_formulas or not isinstance(value, str):
        return value, False
    stripped = value.lstrip(FORMULA_LEADING_WHITESPACE)
    starts_with_control = value.startswith(("\t", "\r", "\n", "\ufeff"))
    if starts_with_control or stripped.startswith(DANGEROUS_FORMULA_PREFIXES):
        return "'" + value, True
    return value, False


def is_dangerous_formula_text(value: Any) -> bool:
    _, changed = sanitize_formula_text(value, allow_formulas=False)
    return changed


def csv_quoting(name: str) -> int:
    values = {
        "minimal": csv.QUOTE_MINIMAL,
        "all": csv.QUOTE_ALL,
        "nonnumeric": csv.QUOTE_NONNUMERIC,
        "none": csv.QUOTE_NONE,
    }
    return values[name]


def rows_for_extract(
    worksheet: Any,
    bounds: tuple[int, int, int, int],
    *,
    sanitize_delimited: bool,
    allow_formulas: bool,
    date_format: str | None,
) -> tuple[list[list[Any]], int]:
    rows: list[list[Any]] = []
    guarded = 0
    for values in iter_range_values(worksheet, bounds):
        output_row = []
        for value in values:
            if date_format and isinstance(value, (dt.datetime, dt.date)):
                value = value.strftime(date_format)
            sanitized, changed = (
                sanitize_formula_text(value, allow_formulas=allow_formulas)
                if sanitize_delimited
                else (value, False)
            )
            output_row.append(serialize_cell_value(sanitized))
            guarded += changed
        rows.append(output_row)
    return rows, guarded


def handle_extract(args: argparse.Namespace) -> dict[str, Any]:
    package = preflight_xlsx(args.source)
    data_only = args.values == "cached"
    workbook = open_workbook(args.source, data_only=data_only)
    worksheet = worksheet_by_name(workbook, args.sheet)
    bounds = normalize_range(worksheet, args.range)
    rows, guarded = rows_for_extract(
        worksheet,
        bounds,
        sanitize_delimited=args.format in {"csv", "tsv"},
        allow_formulas=args.allow_formulas,
        date_format=args.date_format,
    )
    warnings: list[str] = []
    if data_only:
        warnings.append(
            "Cached formula values may be stale or absent because openpyxl does not calculate."
        )
    if guarded:
        warnings.append(
            f"Prefixed {guarded} potentially executable text value(s) with an apostrophe."
        )
    output_format = args.format
    if output_format == "json":

        def writer(handle: Any) -> None:
            json.dump(
                {
                    "schema_version": SCHEMA_VERSION,
                    "source": str(args.source.resolve()),
                    "sheet": worksheet.title,
                    "range": range_to_a1(bounds),
                    "values_mode": args.values,
                    "rows": rows,
                },
                handle,
                ensure_ascii=False,
                sort_keys=True,
                default=json_default,
            )
            handle.write("\n")
    else:
        delimiter = args.delimiter or ("\t" if output_format == "tsv" else ",")
        if len(delimiter) != 1:
            fail("bad_input", "--delimiter must be exactly one character.")

        def writer(handle: Any) -> None:
            csv_writer = csv.writer(
                handle,
                delimiter=delimiter,
                quoting=csv_quoting(args.quoting),
                lineterminator="\n",
                escapechar="\\" if args.quoting == "none" else None,
            )
            csv_writer.writerows(
                [[args.na_value if value is None else value for value in row] for row in rows]
            )

    verification = write_text_atomic(
        args.output,
        writer,
        source=args.source,
        overwrite=args.overwrite,
        encoding=args.encoding,
    )
    workbook.close()
    return success(
        "extract",
        {
            "source": str(args.source.resolve()),
            "output": str(args.output.resolve()),
            "format": output_format,
            "sheet": worksheet.title,
            "range": range_to_a1(bounds),
            "rows": len(rows),
            "columns": max((len(row) for row in rows), default=0),
            "formula_injection_guarded": guarded,
            "package": package,
            "verification": verification,
        },
        warnings,
    )


def unique_sheet_title(workbook: Any, desired: str) -> str:
    require_string(desired, "sheet name")
    if (
        len(desired) > 31
        or re.search(r"[\x00-\x1f\[\]:*?/\\]", desired)
        or desired.startswith("'")
        or desired.endswith("'")
    ):
        fail("bad_input", "Invalid Excel worksheet name.", sheet=desired)
    if desired.casefold() not in {name.casefold() for name in workbook.sheetnames}:
        return desired
    fail("ambiguous_edit", "Worksheet name already exists.", sheet=desired)


def add_named_styles(workbook: Any, specs: Sequence[Any]) -> int:
    existing = {style.name for style in workbook._named_styles}
    count = 0
    for index, raw_spec in enumerate(specs):
        spec = require_mapping(raw_spec, f"named_styles[{index}]")
        name = require_string(spec.get("name"), f"named_styles[{index}].name")
        if name in existing:
            fail("ambiguous_edit", "Named style already exists.", style=name)
        allowed = {
            "name",
            "font",
            "fill",
            "border",
            "alignment",
            "protection",
            "number_format",
        }
        unknown = set(spec) - allowed
        if unknown:
            fail("bad_input", "Unsupported named-style property.", properties=sorted(unknown))
        style = NamedStyle(name=name)
        if "font" in spec:
            style.font = build_font(require_mapping(spec["font"], "named style font"))
        if "fill" in spec:
            style.fill = build_fill(require_mapping(spec["fill"], "named style fill"))
        if "border" in spec:
            style.border = build_border(require_mapping(spec["border"], "named style border"))
        if "alignment" in spec:
            style.alignment = build_alignment(
                require_mapping(spec["alignment"], "named style alignment")
            )
        if "protection" in spec:
            style.protection = build_protection(
                require_mapping(spec["protection"], "named style protection")
            )
        if "number_format" in spec:
            style.number_format = require_string(spec["number_format"], "number_format")
        workbook.add_named_style(style)
        existing.add(name)
        count += 1
    return count


def table_names(workbook: Any) -> set[str]:
    return {
        table.displayName
        for worksheet in workbook.worksheets
        for table in worksheet.tables.values()
    }


def validate_table_name(name: Any, label: str) -> str:
    value = require_string(name, label)
    if not re.fullmatch(r"[A-Za-z_\\][A-Za-z0-9_.\\]*", value) or value.upper() in {
        "R",
        "C",
    }:
        fail("bad_input", "Invalid Excel table name.", table=value)
    return value


def validate_table_range(worksheet: Any, name: str, reference: Any) -> str:
    value = require_string(reference, "table.range")
    try:
        min_col, min_row, max_col, max_row = range_boundaries(value)
    except (TypeError, ValueError) as exc:
        fail("bad_input", "Invalid table range.", range=value, reason=str(exc))
    if not all(isinstance(item, int) for item in (min_col, min_row, max_col, max_row)):
        fail("bad_input", "Table range must contain bounded rows and columns.", range=value)
    check_cell_budget(max_row - min_row + 1, max_col - min_col + 1, label="Table range")
    if max_row <= min_row:
        fail("bad_input", "A table must include a header and at least one data row.")
    headers = [worksheet.cell(min_row, column).value for column in range(min_col, max_col + 1)]
    if any(item is None or str(item) == "" for item in headers):
        fail("bad_input", "Every table column must have a non-empty header.", table=name)
    if len({str(item) for item in headers}) != len(headers):
        fail("bad_input", "Table headers must be unique.", table=name)
    return value


def build_table_style(spec: Mapping[str, Any]) -> Any:
    reject_unknown_keys(
        spec,
        {
            "name",
            "show_first_column",
            "show_last_column",
            "show_row_stripes",
            "show_column_stripes",
        },
        "table.style",
    )
    return TableStyleInfo(
        name=spec.get("name", "TableStyleMedium2"),
        showFirstColumn=bool(spec.get("show_first_column", False)),
        showLastColumn=bool(spec.get("show_last_column", False)),
        showRowStripes=bool(spec.get("show_row_stripes", True)),
        showColumnStripes=bool(spec.get("show_column_stripes", False)),
    )


def ranges_overlap(first: str, second: str) -> bool:
    a_min_col, a_min_row, a_max_col, a_max_row = range_boundaries(first)
    b_min_col, b_min_row, b_max_col, b_max_row = range_boundaries(second)
    return not (
        a_max_col < b_min_col
        or b_max_col < a_min_col
        or a_max_row < b_min_row
        or b_max_row < a_min_row
    )


def add_table(worksheet: Any, spec: Mapping[str, Any], workbook: Any) -> Any:
    reject_unknown_keys(
        spec,
        {"name", "range", "style", "totals_row_shown"},
        "table",
    )
    name = validate_table_name(spec.get("name"), "table.name")
    if name in table_names(workbook):
        fail("ambiguous_edit", "Table name already exists in the workbook.", table=name)
    reference = validate_table_range(worksheet, name, spec.get("range"))
    style = build_table_style(require_mapping(spec.get("style", {}), "table.style"))
    table = Table(displayName=name, ref=reference)
    table.tableStyleInfo = style
    table.totalsRowShown = bool(spec.get("totals_row_shown", False))
    # Excel treats a sheet-level auto filter overlapping a table as invalid content
    # and offers document recovery; the table supplies its own filter.
    existing_filter = worksheet.auto_filter.ref
    if existing_filter and ranges_overlap(existing_filter, reference):
        worksheet.auto_filter.ref = None
    worksheet.add_table(table)
    return table


def color_list(values: Any, label: str) -> list[str]:
    return [
        color_value(value, f"{label}[{index}]")
        for index, value in enumerate(require_list(values, label))
    ]


def add_conditional_format(worksheet: Any, spec: Mapping[str, Any]) -> None:
    reject_unknown_keys(
        spec,
        {
            "range",
            "type",
            "style",
            "formula",
            "operator",
            "stop_if_true",
            "colors",
            "value_types",
            "values",
            "start_type",
            "start_value",
            "end_type",
            "end_value",
            "color",
            "show_value",
            "icon_style",
            "value_type",
            "reverse",
        },
        "conditional format",
    )
    reference = require_string(spec.get("range"), "conditional_format.range")
    rule_type = require_string(spec.get("type"), "conditional_format.type")
    style_spec = require_mapping(spec.get("style", {}), "conditional_format.style")
    fill = (
        build_fill(require_mapping(style_spec["fill"], "style.fill"))
        if "fill" in style_spec
        else None
    )
    font = (
        build_font(require_mapping(style_spec["font"], "style.font"))
        if "font" in style_spec
        else None
    )
    border = (
        build_border(require_mapping(style_spec["border"], "style.border"))
        if "border" in style_spec
        else None
    )
    if set(style_spec) - {"fill", "font", "border"}:
        fail("bad_input", "Unsupported conditional-format style property.")
    if rule_type == "cell_is":
        formula = [str(value) for value in require_list(spec.get("formula", []), "formula")]
        rule = CellIsRule(
            operator=require_string(spec.get("operator"), "operator"),
            formula=formula,
            stopIfTrue=bool(spec.get("stop_if_true", False)),
            font=font,
            border=border,
            fill=fill,
        )
    elif rule_type == "formula":
        formula = [str(value) for value in require_list(spec.get("formula", []), "formula")]
        if not formula:
            fail("bad_input", "Formula conditional formatting requires formula entries.")
        rule = FormulaRule(
            formula=formula,
            stopIfTrue=bool(spec.get("stop_if_true", False)),
            font=font,
            border=border,
            fill=fill,
        )
    elif rule_type == "color_scale":
        colors = color_list(spec.get("colors"), "conditional_format.colors")
        if len(colors) not in {2, 3}:
            fail("bad_input", "Color scales require exactly two or three colors.")
        default_types = ["min", "max"] if len(colors) == 2 else ["min", "percentile", "max"]
        types = spec.get("value_types", default_types)
        values = spec.get("values", [None, None] if len(colors) == 2 else [None, 50, None])
        if len(types) != len(colors) or len(values) != len(colors):
            fail("bad_input", "Color-scale types, values, and colors must have equal lengths.")
        kwargs: dict[str, Any] = {}
        for index, suffix in enumerate(("start", "mid", "end")[: len(colors)]):
            kwargs[f"{suffix}_type"] = types[index]
            kwargs[f"{suffix}_value"] = values[index]
            kwargs[f"{suffix}_color"] = colors[index]
        rule = ColorScaleRule(**kwargs)
    elif rule_type == "data_bar":
        rule = DataBarRule(
            start_type=spec.get("start_type", "min"),
            start_value=spec.get("start_value"),
            end_type=spec.get("end_type", "max"),
            end_value=spec.get("end_value"),
            color=color_value(spec.get("color", "638EC6"), "conditional_format.color"),
            showValue=bool(spec.get("show_value", True)),
        )
    elif rule_type == "icon_set":
        rule = IconSetRule(
            icon_style=spec.get("icon_style", "3TrafficLights1"),
            type=spec.get("value_type", "percent"),
            values=require_list(spec.get("values", [0, 33, 67]), "values"),
            showValue=bool(spec.get("show_value", True)),
            reverse=bool(spec.get("reverse", False)),
        )
    else:
        fail(
            "unsupported_operation",
            "Unsupported conditional-format rule type.",
            type=rule_type,
        )
    worksheet.conditional_formatting.add(reference, rule)


def add_validation(worksheet: Any, spec: Mapping[str, Any]) -> None:
    reject_unknown_keys(
        spec,
        {
            "range",
            "type",
            "formula1",
            "formula2",
            "operator",
            "allow_blank",
            "show_dropdown",
            "show_error_message",
            "show_input_message",
            "error_title",
            "error",
            "prompt_title",
            "prompt",
        },
        "data validation",
    )
    reference = require_string(spec.get("range"), "data_validation.range")
    validation_type = require_string(spec.get("type"), "data_validation.type")
    validation = DataValidation(
        type=validation_type,
        formula1=spec.get("formula1"),
        formula2=spec.get("formula2"),
        operator=spec.get("operator"),
        allow_blank=bool(spec.get("allow_blank", False)),
        showDropDown=spec.get("show_dropdown"),
        showErrorMessage=bool(spec.get("show_error_message", True)),
        showInputMessage=bool(spec.get("show_input_message", True)),
        errorTitle=spec.get("error_title"),
        error=spec.get("error"),
        promptTitle=spec.get("prompt_title"),
        prompt=spec.get("prompt"),
    )
    worksheet.add_data_validation(validation)
    validation.add(reference)


def parse_sheet_range(workbook: Any, default_sheet: Any, raw: str, label: str) -> tuple[Any, str]:
    text = require_string(raw, label)
    if "!" not in text:
        return default_sheet, text
    sheet_part, reference = text.rsplit("!", 1)
    sheet_name = sheet_part.strip("'").replace("''", "'")
    if sheet_name not in workbook.sheetnames:
        fail("bad_input", f"{label} references an unknown sheet.", sheet=sheet_name)
    return workbook[sheet_name], reference


def range_boundaries_for_user(reference: str, label: str) -> tuple[int, int, int, int]:
    try:
        bounds = range_boundaries(reference)
    except (TypeError, ValueError) as exc:
        fail("bad_input", f"{label} is not a valid A1 range.", range=reference, reason=str(exc))
    if not all(isinstance(value, int) for value in bounds):
        fail("bad_input", f"{label} must contain bounded rows and columns.", range=reference)
    return bounds


def add_chart(worksheet: Any, spec: Mapping[str, Any], workbook: Any) -> Any:
    reject_unknown_keys(
        spec,
        {
            "type",
            "title",
            "style",
            "height",
            "width",
            "x_axis_title",
            "y_axis_title",
            "x_values",
            "y_values",
            "series_titles",
            "data",
            "titles_from_data",
            "from_rows",
            "categories",
            "bar_direction",
            "grouping",
            "anchor",
        },
        "chart",
    )
    chart_type = require_string(spec.get("type"), "chart.type").lower()
    chart_classes = {
        "bar": BarChart,
        "line": LineChart,
        "pie": PieChart,
        "area": AreaChart,
        "scatter": ScatterChart,
    }
    if chart_type not in chart_classes:
        fail("unsupported_operation", "Unsupported native chart type.", type=chart_type)
    chart = chart_classes[chart_type]()
    if chart_type == "scatter":
        # ECMA-376 requires CT_ScatterChart/scatterStyle; openpyxl omits it when unset
        # and Excel then offers document recovery.
        chart.scatterStyle = "lineMarker"
    chart.title = spec.get("title")
    chart.style = spec.get("style", 10)
    chart.height = spec.get("height", 7.5)
    chart.width = spec.get("width", 12.5)
    if "x_axis_title" in spec:
        chart.x_axis.title = spec["x_axis_title"]
    if "y_axis_title" in spec:
        chart.y_axis.title = spec["y_axis_title"]
    if chart_type == "scatter":
        x_sheet, x_ref = parse_sheet_range(
            workbook,
            worksheet,
            require_string(spec.get("x_values"), "chart.x_values"),
            "chart.x_values",
        )
        x_min_col, x_min_row, x_max_col, x_max_row = range_boundaries_for_user(
            x_ref, "chart.x_values"
        )
        if x_min_col != x_max_col:
            fail("bad_input", "Scatter chart x_values must be one column.")
        x_values = Reference(
            x_sheet,
            min_col=x_min_col,
            min_row=x_min_row,
            max_row=x_max_row,
        )
        y_ranges = require_list(spec.get("y_values"), "chart.y_values")
        for index, raw_range in enumerate(y_ranges):
            y_sheet, y_ref = parse_sheet_range(
                workbook, worksheet, require_string(raw_range, "y range"), "chart.y_values"
            )
            y_min_col, y_min_row, y_max_col, y_max_row = range_boundaries_for_user(
                y_ref, "chart.y_values"
            )
            if y_min_col != y_max_col or y_max_row - y_min_row != x_max_row - x_min_row:
                fail("bad_input", "Scatter y ranges must match the x range dimensions.")
            y_values = Reference(
                y_sheet,
                min_col=y_min_col,
                min_row=y_min_row,
                max_row=y_max_row,
            )
            title = None
            titles = spec.get("series_titles", [])
            if index < len(titles):
                title = str(titles[index])
            chart.series.append(Series(y_values, x_values, title=title))
    else:
        data_sheet, data_ref = parse_sheet_range(
            workbook,
            worksheet,
            require_string(spec.get("data"), "chart.data"),
            "chart.data",
        )
        min_col, min_row, max_col, max_row = range_boundaries_for_user(data_ref, "chart.data")
        data = Reference(
            data_sheet,
            min_col=min_col,
            min_row=min_row,
            max_col=max_col,
            max_row=max_row,
        )
        chart.add_data(
            data,
            titles_from_data=bool(spec.get("titles_from_data", True)),
            from_rows=bool(spec.get("from_rows", False)),
        )
        if "categories" in spec:
            cat_sheet, cat_ref = parse_sheet_range(
                workbook,
                worksheet,
                require_string(spec["categories"], "chart.categories"),
                "chart.categories",
            )
            cat_min_col, cat_min_row, cat_max_col, cat_max_row = range_boundaries_for_user(
                cat_ref, "chart.categories"
            )
            categories = Reference(
                cat_sheet,
                min_col=cat_min_col,
                min_row=cat_min_row,
                max_col=cat_max_col,
                max_row=cat_max_row,
            )
            chart.set_categories(categories)
        if chart_type == "bar":
            chart.type = spec.get("bar_direction", "col")
            chart.grouping = spec.get("grouping", "clustered")
        elif hasattr(chart, "grouping") and "grouping" in spec:
            chart.grouping = spec["grouping"]
    anchor = require_string(spec.get("anchor", "E2"), "chart.anchor")
    worksheet.add_chart(chart, anchor)
    return chart


def set_sheet_dimensions(worksheet: Any, spec: Mapping[str, Any]) -> None:
    reject_unknown_keys(spec, {"columns", "rows"}, "dimensions")
    columns = require_mapping(spec.get("columns", {}), "dimensions.columns")
    for column, raw in columns.items():
        data = raw if isinstance(raw, dict) else {"width": raw}
        reject_unknown_keys(data, {"width", "hidden", "outline_level"}, "column dimension")
        dimension = worksheet.column_dimensions[column]
        if "width" in data:
            dimension.width = float(data["width"])
        if "hidden" in data:
            dimension.hidden = bool(data["hidden"])
        if "outline_level" in data:
            dimension.outlineLevel = int(data["outline_level"])
    rows = require_mapping(spec.get("rows", {}), "dimensions.rows")
    for row_text, raw in rows.items():
        try:
            row = int(row_text)
        except (TypeError, ValueError):
            fail("bad_input", "Row dimension keys must be integers.", row=row_text)
        data = raw if isinstance(raw, dict) else {"height": raw}
        reject_unknown_keys(data, {"height", "hidden", "outline_level"}, "row dimension")
        dimension = worksheet.row_dimensions[row]
        if "height" in data:
            dimension.height = float(data["height"])
        if "hidden" in data:
            dimension.hidden = bool(data["hidden"])
        if "outline_level" in data:
            dimension.outlineLevel = int(data["outline_level"])


def schema_convert_series(series: Any, type_name: str, *, errors: str = "raise") -> Any:
    if errors not in {"raise", "coerce"}:
        fail("bad_input", "Conversion errors policy must be 'raise' or 'coerce'.")
    if type_name == "string":
        return series.astype("string")
    if type_name == "integer":
        converted = pd.to_numeric(series, errors=errors)
        return converted.astype("Int64")
    if type_name == "float":
        return pd.to_numeric(series, errors=errors)
    if type_name == "boolean":
        lowered = series.astype("string").str.strip().str.lower()
        mapping = {
            "true": True,
            "false": False,
            "1": True,
            "0": False,
            "yes": True,
            "no": False,
        }
        converted = lowered.map(mapping)
        invalid = converted.isna() & series.notna()
        if errors == "raise" and invalid.any():
            bad = series[invalid].head(5).tolist()
            fail("bad_input", "Could not coerce values to boolean.", examples=bad)
        return converted.astype("boolean")
    if type_name in {"date", "datetime"}:
        converted = pd.to_datetime(series, errors=errors)
        return converted.dt.date if type_name == "date" else converted
    fail("bad_input", "Unsupported schema type.", type=type_name)


def apply_dataframe_schema(
    frame: Any,
    schema: Mapping[str, Any],
    *,
    default_errors: str = "raise",
) -> Any:
    result = frame.copy()
    for column, raw_spec in schema.items():
        if column not in result.columns:
            fail("bad_input", "Schema references an unknown column.", column=column)
        if isinstance(raw_spec, str):
            type_name = raw_spec
            errors = default_errors
        else:
            spec = require_mapping(raw_spec, f"schema.{column}")
            reject_unknown_keys(spec, {"type", "errors"}, f"schema.{column}")
            type_name = require_string(spec.get("type"), f"schema.{column}.type")
            errors = spec.get("errors", default_errors)
        try:
            result[column] = schema_convert_series(result[column], type_name, errors=errors)
        except ToolError:
            raise
        except (TypeError, ValueError, OverflowError) as exc:
            fail(
                "bad_input",
                "Could not apply column schema.",
                column=column,
                type=type_name,
                reason=str(exc),
            )
    return result


def read_delimited_frame(
    source_spec: Mapping[str, Any],
    *,
    base_directory: Path,
) -> tuple[Any, dict[str, Any]]:
    reject_unknown_keys(
        source_spec,
        {
            "path",
            "format",
            "encoding",
            "delimiter",
            "header",
            "leading_zeros",
            "recognize_default_na",
            "na_values",
            "malformed_rows",
            "quoting",
            "schema",
        },
        "delimited source",
    )
    raw_path = Path(require_string(source_spec.get("path"), "source.path"))
    path = raw_path if raw_path.is_absolute() else base_directory / raw_path
    format_name = source_spec.get("format", path.suffix.lower().lstrip("."))
    if format_name not in {"csv", "tsv"}:
        fail("bad_input", "Delimited source format must be csv or tsv.", format=format_name)
    encoding = source_spec.get("encoding", "utf-8")
    delimiter = source_spec.get("delimiter", "\t" if format_name == "tsv" else ",")
    if not isinstance(delimiter, str) or len(delimiter) != 1:
        fail("bad_input", "Delimited source delimiter must be one character.")
    if not path.is_file():
        fail("bad_input", "Delimited source does not exist.", path=path)
    leading_zeros = source_spec.get("leading_zeros", "preserve")
    if leading_zeros not in {"preserve", "numeric"}:
        fail("bad_input", "leading_zeros must be 'preserve' or 'numeric'.")
    dtype = str if leading_zeros == "preserve" else None
    try:
        frame = pd.read_csv(
            path,
            sep=delimiter,
            encoding=encoding,
            header=0 if source_spec.get("header", True) else None,
            dtype=dtype,
            keep_default_na=bool(source_spec.get("recognize_default_na", False)),
            na_values=source_spec.get("na_values"),
            on_bad_lines=source_spec.get("malformed_rows", "error"),
            quoting=csv_quoting(source_spec.get("quoting", "minimal")),
        )
    except (LookupError, OSError, UnicodeError, pd.errors.ParserError, ValueError) as exc:
        fail("bad_input", "Could not parse delimited source.", path=path, reason=str(exc))
    header_cells = len(frame.columns) if source_spec.get("header", True) else 0
    check_cell_budget(
        len(frame) * len(frame.columns) + header_cells,
        1,
        label="Delimited source",
    )
    schema = require_mapping(source_spec.get("schema", {}), "source.schema")
    if schema:
        frame = apply_dataframe_schema(frame, schema)
    return frame, {
        "path": str(path.resolve()),
        "format": format_name,
        "encoding": encoding,
        "delimiter": delimiter,
        "header": bool(source_spec.get("header", True)),
        "recognize_default_na": bool(source_spec.get("recognize_default_na", False)),
        "na_values": source_spec.get("na_values"),
        "rows": len(frame),
        "columns": len(frame.columns),
    }


def dataframe_rows(frame: Any, *, include_header: bool = True) -> list[list[Any]]:
    rows: list[list[Any]] = []
    if include_header:
        rows.append([str(column) for column in frame.columns])
    for raw_row in frame.itertuples(index=False, name=None):
        row: list[Any] = []
        for value in raw_row:
            if pd.isna(value):
                row.append(None)
            elif isinstance(value, pd.Timestamp):
                row.append(value.to_pydatetime())
            elif isinstance(value, np.generic):
                row.append(value.item())
            else:
                row.append(value)
        rows.append(row)
    return rows


def populate_rows(
    worksheet: Any,
    rows: Sequence[Any],
    *,
    start_row: int = 1,
    budget: CellBudget | None = None,
) -> int:
    validated: list[list[Any]] = []
    for row_offset, raw_row in enumerate(rows):
        validated.append(require_list(raw_row, f"rows[{row_offset}]"))
    written = sum(len(row) for row in validated)
    if budget is not None:
        budget.consume(written, label="JSON rows")
    for row_offset, row in enumerate(validated):
        for column, value in enumerate(row, start=1):
            set_cell_from_spec(worksheet.cell(start_row + row_offset, column), value)
    return written


MERGE_SCAN_CELL_LIMIT = 100_000


def merge_covered_values(worksheet: Any, range_text: str) -> list[str]:
    """Coordinates of value-bearing cells a merge would discard (all but top-left)."""

    try:
        min_col, min_row, max_col, max_row = range_boundaries(range_text)
    except ValueError as exc:
        fail("bad_input", "Invalid merge range.", range=range_text, reason=str(exc))
    if None in (min_col, min_row, max_col, max_row):
        fail("bad_input", "Merge ranges must be bounded A1 rectangles.", range=range_text)
    span = (max_row - min_row + 1) * (max_col - min_col + 1)
    if span > MERGE_SCAN_CELL_LIMIT:
        fail(
            "resource_limit",
            "Merge range exceeds the data-loss scan limit.",
            range=range_text,
            cells=span,
            limit=MERGE_SCAN_CELL_LIMIT,
        )
    occupied: list[str] = []
    for row in worksheet.iter_rows(
        min_row=min_row,
        min_col=min_col,
        max_row=max_row,
        max_col=max_col,
    ):
        for cell in row:
            if cell.row == min_row and cell.column == min_col:
                continue
            if cell.value is not None:
                occupied.append(cell.coordinate)
    return occupied


def parse_auto_width_spec(value: Any, label: str) -> dict[str, Any] | None:
    if value is None or value is False:
        return None
    if value is True:
        return dict(AUTO_WIDTH_DEFAULTS)
    spec = require_mapping(value, label)
    reject_unknown_keys(spec, {"columns", "min_width", "max_width", "sample_rows"}, label)
    parsed: dict[str, Any] = dict(AUTO_WIDTH_DEFAULTS)
    if "columns" in spec:
        columns = require_list(spec["columns"], f"{label}.columns")
        if not columns:
            fail("bad_input", f"{label}.columns must not be empty.")
        parsed["columns"] = [
            require_string(column, f"{label} column").upper() for column in columns
        ]
    for key in ("min_width", "max_width", "sample_rows"):
        if key in spec:
            bound = spec[key]
            if isinstance(bound, bool) or not isinstance(bound, int) or bound < 1:
                fail("bad_input", f"{label}.{key} must be a positive integer.")
            parsed[key] = bound
    if parsed["min_width"] > parsed["max_width"]:
        fail("bad_input", f"{label}.min_width cannot exceed max_width.")
    return parsed


def auto_width_value(sample_lengths: Sequence[int], *, min_width: int, max_width: int) -> float:
    widest = max(sample_lengths, default=0)
    return float(min(max(widest, min_width) + 2, max_width))


def apply_auto_width(worksheet: Any, spec: Mapping[str, Any]) -> int:
    """Set character-count-heuristic column widths from stored cell text."""

    if "columns" in spec:
        try:
            targets = [column_index_from_string(letter) for letter in spec["columns"]]
        except ValueError as exc:
            fail(
                "bad_input",
                "auto_width.columns must contain column letters.",
                reason=str(exc),
            )
    else:
        targets = list(range(1, (worksheet.max_column or 1) + 1))
    sample_rows = min(int(spec["sample_rows"]), worksheet.max_row or 1)
    for column in targets:
        lengths = [
            len(str(cell.value))
            for row in worksheet.iter_rows(
                min_row=1,
                max_row=sample_rows,
                min_col=column,
                max_col=column,
            )
            for cell in row
            if cell.value is not None
        ]
        worksheet.column_dimensions[get_column_letter(column)].width = auto_width_value(
            lengths,
            min_width=spec["min_width"],
            max_width=spec["max_width"],
        )
    return len(targets)


def apply_sheet_settings(
    worksheet: Any,
    spec: Mapping[str, Any],
    *,
    context: str,
    warnings: list[str],
) -> Counter[str]:
    """Apply the worksheet settings shared by create sheets, add_sheet, and set_sheet."""

    counts: Counter[str] = Counter()
    allow_merge_data_loss = bool(spec.get("allow_merge_data_loss", False))
    for reference in require_list(spec.get("merges", []), "sheet.merges"):
        range_text = require_string(reference, "merge range").upper()
        occupied = merge_covered_values(worksheet, range_text)
        if occupied:
            if context == "edit" and not allow_merge_data_loss:
                fail(
                    "ambiguous_edit",
                    "Merging this range would discard non-top-left cell values.",
                    range=range_text,
                    sheet=worksheet.title,
                    occupied_cells=occupied[:10],
                    occupied_count=len(occupied),
                    resolution="Set allow_merge_data_loss to true to accept the loss.",
                )
            warnings.append(
                f"Merging {range_text} on sheet {worksheet.title!r} discarded "
                f"{len(occupied)} non-top-left cell value(s); only the top-left value is kept."
            )
        worksheet.merge_cells(range_text)
        counts["ranges_merged"] += 1
    for reference in require_list(spec.get("unmerge", []), "sheet.unmerge"):
        range_text = require_string(reference, "unmerge range").upper()
        existing = sorted(str(item) for item in worksheet.merged_cells.ranges)
        if range_text not in existing:
            fail(
                "bad_input",
                "unmerge must exactly match an existing merged range.",
                range=range_text,
                sheet=worksheet.title,
                merged=existing,
            )
        worksheet.unmerge_cells(range_text)
        counts["ranges_unmerged"] += 1
    if "freeze_panes" in spec:
        worksheet.freeze_panes = spec["freeze_panes"]
    if "auto_filter" in spec:
        worksheet.auto_filter.ref = require_string(spec["auto_filter"], "auto_filter")
    if "tab_color" in spec:
        worksheet.sheet_properties.tabColor = color_value(spec["tab_color"], "tab_color")
    if "show_gridlines" in spec:
        worksheet.sheet_view.showGridLines = bool(spec["show_gridlines"])
    if "zoom_scale" in spec:
        zoom = int(spec["zoom_scale"])
        if not 10 <= zoom <= 400:
            fail("bad_input", "zoom_scale must be between 10 and 400.")
        worksheet.sheet_view.zoomScale = zoom
    auto_width = parse_auto_width_spec(spec.get("auto_width"), "auto_width")
    if auto_width is not None:
        counts["columns_auto_sized"] += apply_auto_width(worksheet, auto_width)
    if "dimensions" in spec:
        set_sheet_dimensions(
            worksheet,
            require_mapping(spec["dimensions"], "sheet.dimensions"),
        )
    return counts


def configure_worksheet(
    workbook: Any,
    worksheet: Any,
    spec: Mapping[str, Any],
    *,
    base_directory: Path,
    budget: CellBudget,
) -> tuple[dict[str, int], list[str]]:
    reject_unknown_keys(
        spec,
        {
            "op",
            "index",
            "name",
            "source",
            "rows",
            "cells",
            "state",
            "merges",
            "freeze_panes",
            "dimensions",
            "auto_filter",
            "auto_width",
            "tab_color",
            "show_gridlines",
            "zoom_scale",
            "tables",
            "conditional_formats",
            "data_validations",
            "charts",
        },
        "sheet",
    )
    counts: Counter[str] = Counter(
        {
            "cells": 0,
            "tables": 0,
            "charts": 0,
            "conditional_formats": 0,
            "data_validations": 0,
        }
    )
    warnings: list[str] = []
    if "source" in spec and "rows" in spec:
        fail("bad_input", "A sheet cannot contain both source and rows.")
    if "source" in spec:
        source_spec = require_mapping(spec["source"], "sheet.source")
        frame, _ = read_delimited_frame(
            source_spec,
            base_directory=base_directory,
        )
        counts["cells"] += populate_rows(
            worksheet,
            dataframe_rows(frame, include_header=bool(source_spec.get("header", True))),
            budget=budget,
        )
    elif "rows" in spec:
        counts["cells"] += populate_rows(
            worksheet,
            require_list(spec["rows"], "sheet.rows"),
            budget=budget,
        )
    cells = require_mapping(spec.get("cells", {}), "sheet.cells")
    budget.consume(len(cells), label="JSON cell mappings")
    for coordinate, cell_spec in cells.items():
        if not isinstance(coordinate, str):
            fail("bad_input", "Cell coordinates must be strings.")
        try:
            cell = worksheet[coordinate]
        except (KeyError, ValueError) as exc:
            fail("bad_input", "Invalid cell coordinate.", coordinate=coordinate, reason=str(exc))
        if not hasattr(cell, "coordinate"):
            fail("bad_input", "Cell mapping keys must identify one cell.", coordinate=coordinate)
        set_cell_from_spec(cell, cell_spec)
        counts["cells"] += 1
    counts.update(
        apply_sheet_settings(
            worksheet,
            spec,
            context="create",
            warnings=warnings,
        )
    )
    for raw_table in require_list(spec.get("tables", []), "sheet.tables"):
        add_table(worksheet, require_mapping(raw_table, "table"), workbook)
        counts["tables"] += 1
    for raw_rule in require_list(spec.get("conditional_formats", []), "sheet.conditional_formats"):
        add_conditional_format(
            worksheet,
            require_mapping(raw_rule, "conditional format"),
        )
        counts["conditional_formats"] += 1
    for raw_validation in require_list(spec.get("data_validations", []), "sheet.data_validations"):
        add_validation(
            worksheet,
            require_mapping(raw_validation, "data validation"),
        )
        counts["data_validations"] += 1
    for raw_chart in require_list(spec.get("charts", []), "sheet.charts"):
        add_chart(worksheet, require_mapping(raw_chart, "chart"), workbook)
        counts["charts"] += 1
    state = spec.get("state", "visible")
    if state not in {"visible", "hidden", "veryHidden"}:
        fail("bad_input", "Invalid worksheet state.", state=state)
    worksheet.sheet_state = state
    return dict(counts), warnings


def set_workbook_properties(workbook: Any, properties: Mapping[str, Any]) -> None:
    allowed = {
        "title",
        "subject",
        "creator",
        "keywords",
        "description",
        "category",
        "last_modified_by",
        "created",
        "modified",
    }
    unknown = set(properties) - allowed
    if unknown:
        fail("bad_input", "Unsupported workbook property.", properties=sorted(unknown))
    aliases = {"last_modified_by": "lastModifiedBy"}
    for key, value in properties.items():
        attribute = aliases.get(key, key)
        if key in {"created", "modified"} and isinstance(value, str):
            try:
                value = dt.datetime.fromisoformat(value.replace("Z", "+00:00"))
            except ValueError as exc:
                fail("bad_input", f"{key} must be an ISO-8601 date-time.", reason=str(exc))
        setattr(workbook.properties, attribute, value)


def add_defined_names(workbook: Any, specs: Sequence[Any]) -> int:
    count = 0
    for index, raw_spec in enumerate(specs):
        spec = require_mapping(raw_spec, f"defined_names[{index}]")
        reject_unknown_keys(
            spec,
            {"name", "refers_to", "local_sheet", "hidden"},
            f"defined_names[{index}]",
        )
        name = require_string(spec.get("name"), "defined name")
        if name in workbook.defined_names:
            fail("ambiguous_edit", "Defined name already exists.", name=name)
        text = require_string(spec.get("refers_to"), "defined_name.refers_to")
        local_sheet = spec.get("local_sheet")
        local_sheet_id = None
        if local_sheet is not None:
            if local_sheet not in workbook.sheetnames:
                fail("bad_input", "Defined name local_sheet does not exist.", sheet=local_sheet)
            local_sheet_id = workbook.sheetnames.index(local_sheet)
        defined = DefinedName(
            name,
            attr_text=text,
            localSheetId=local_sheet_id,
            hidden=bool(spec.get("hidden", False)),
        )
        workbook.defined_names.add(defined)
        count += 1
    return count


def handle_create(args: argparse.Namespace) -> dict[str, Any]:
    job = load_json_object(args.job, "create job")
    validate_job(job)
    reject_unknown_keys(
        job,
        {
            "schema_version",
            "properties",
            "named_styles",
            "sheets",
            "defined_names",
            "calculation",
        },
        "create job",
    )
    workbook = Workbook()
    workbook.remove(workbook.active)
    named_style_count = add_named_styles(
        workbook,
        require_list(job.get("named_styles", []), "named_styles"),
    )
    sheets = require_list(job.get("sheets"), "sheets")
    if not sheets:
        fail("bad_input", "Create job must contain at least one sheet.")
    total_counts: Counter[str] = Counter()
    budget = cell_budget_from_args(args)
    sheet_warnings: list[str] = []
    for index, raw_spec in enumerate(sheets):
        spec = require_mapping(raw_spec, f"sheets[{index}]")
        name = unique_sheet_title(
            workbook,
            require_string(spec.get("name"), f"sheets[{index}].name"),
        )
        worksheet = workbook.create_sheet(name)
        sheet_counts, configure_warnings = configure_worksheet(
            workbook,
            worksheet,
            spec,
            base_directory=args.job.resolve().parent,
            budget=budget,
        )
        total_counts.update(sheet_counts)
        sheet_warnings.extend(configure_warnings)
    if not any(sheet.sheet_state == "visible" for sheet in workbook.worksheets):
        fail("bad_input", "At least one worksheet must remain visible.")
    properties = require_mapping(job.get("properties", {}), "properties")
    set_workbook_properties(workbook, properties)
    defined_name_count = add_defined_names(
        workbook,
        require_list(job.get("defined_names", []), "defined_names"),
    )
    calculation = require_mapping(job.get("calculation", {}), "calculation")
    if calculation:
        reject_unknown_keys(
            calculation,
            {"mode", "full_calc_on_load", "force_full_calc"},
            "calculation",
        )
        mode = calculation.get("mode", "auto")
        if mode not in {"auto", "manual", "autoNoTable"}:
            fail("bad_input", "Unsupported calculation mode.", mode=mode)
        workbook.calculation.calcMode = mode
        workbook.calculation.fullCalcOnLoad = bool(calculation.get("full_calc_on_load", True))
        workbook.calculation.forceFullCalc = bool(calculation.get("force_full_calc", True))
    verification = save_workbook_atomic(
        workbook,
        args.output,
        source=None,
        overwrite=args.overwrite,
    )
    formula_count = sum(
        cell.data_type == "f"
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
    )
    workbook.close()
    warnings = list(sheet_warnings)
    if formula_count:
        warnings.append(
            "Formulas were written but not calculated; calculation flags request, "
            "not prove, refresh."
        )
    return success(
        "create",
        {
            "job": str(args.job.resolve()),
            "output": str(args.output.resolve()),
            "sheets": verification["sheet_names"],
            "counts": {
                **dict(total_counts),
                "named_styles": named_style_count,
                "defined_names": defined_name_count,
                "budgeted_cells": budget.used,
            },
            "verification": verification,
        },
        warnings,
    )


def sheet_reference_tokens(sheet_name: str) -> tuple[str, str]:
    quoted = quote_sheetname(sheet_name) + "!"
    return quoted.casefold(), f"{sheet_name}!".casefold()


def formula_dependencies(workbook: Any, sheet_name: str) -> list[dict[str, Any]]:
    tokens = sheet_reference_tokens(sheet_name)
    dependencies: list[dict[str, Any]] = []
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                text = str(cell.value).casefold()
                if cell.data_type == "f" and any(token in text for token in tokens):
                    dependencies.append(
                        {
                            "kind": "formula",
                            "sheet": worksheet.title,
                            "coordinate": cell.coordinate,
                        }
                    )
    for name, defined in workbook.defined_names.items():
        text = (getattr(defined, "attr_text", "") or "").casefold()
        if any(token in text for token in tokens):
            dependencies.append({"kind": "defined_name", "name": name})
    for worksheet in workbook.worksheets:
        for index, chart in enumerate(getattr(worksheet, "_charts", [])):
            for series in getattr(chart, "series", []):
                text = str(series).casefold()
                if any(token in text for token in tokens):
                    dependencies.append({"kind": "chart", "sheet": worksheet.title, "index": index})
    return dependencies


def require_safe_sheet_change(
    workbook: Any,
    sheet_name: str,
    *,
    acknowledged: bool,
    operation: str,
) -> list[str]:
    dependencies = formula_dependencies(workbook, sheet_name)
    if dependencies and not acknowledged:
        fail(
            "ambiguous_edit",
            f"{operation} could leave formulas, names, or charts unresolved.",
            sheet=sheet_name,
            dependencies=dependencies[:50],
            dependency_count=len(dependencies),
            acknowledgement=(
                "Set allow_unupdated_dependencies to true only after accepting dangling or "
                "unchanged references."
            ),
        )
    return (
        [
            f"{operation} was allowed despite {len(dependencies)} dependency reference(s); "
            "references were not rewritten."
        ]
        if dependencies
        else []
    )


def set_range_values(
    worksheet: Any,
    reference: str,
    values: Any,
    *,
    budget: CellBudget,
) -> int:
    bounds = normalize_range(worksheet, reference)
    rows = require_list(values, "set_range.values")
    min_row, min_col, max_row, max_col = bounds
    expected_rows = max_row - min_row + 1
    expected_columns = max_col - min_col + 1
    if len(rows) != expected_rows:
        fail(
            "bad_input",
            "set_range row count does not match the destination range.",
            expected=expected_rows,
            actual=len(rows),
        )
    for index, raw_row in enumerate(rows):
        row = require_list(raw_row, f"set_range.values[{index}]")
        if len(row) != expected_columns:
            fail(
                "bad_input",
                "set_range column count does not match the destination range.",
                row=index,
                expected=expected_columns,
                actual=len(row),
            )
    budget.consume(expected_rows * expected_columns, label="set_range JSON values")
    for index, raw_row in enumerate(rows):
        row = require_list(raw_row, f"set_range.values[{index}]")
        for offset, value in enumerate(row):
            set_cell_from_spec(worksheet.cell(min_row + index, min_col + offset), value)
    return expected_rows * expected_columns


def format_range(
    worksheet: Any,
    reference: str,
    style_spec: Mapping[str, Any],
    *,
    budget: CellBudget,
) -> int:
    bounds = normalize_range(worksheet, reference)
    min_row, min_col, max_row, max_col = bounds
    budget.consume(
        (max_row - min_row + 1) * (max_col - min_col + 1),
        label="format_range",
    )
    count = 0
    for row in worksheet.iter_rows(
        min_row=min_row,
        min_col=min_col,
        max_row=max_row,
        max_col=max_col,
    ):
        for cell in row:
            apply_style(cell, style_spec)
            count += 1
    return count


def find_table(workbook: Any, name: str) -> tuple[Any, Any]:
    matches = [
        (worksheet, table)
        for worksheet in workbook.worksheets
        for table in worksheet.tables.values()
        if table.displayName == name
    ]
    if not matches:
        fail("bad_input", "Table does not exist.", table=name)
    if len(matches) > 1:
        fail("ambiguous_edit", "Table name is not unique.", table=name)
    return matches[0]


def update_table(workbook: Any, spec: Mapping[str, Any]) -> None:
    reject_unknown_keys(
        spec,
        {"name", "range", "new_name", "style", "totals_row_shown"},
        "update_table.table",
    )
    name = require_string(spec.get("name"), "update_table.name")
    worksheet, table = find_table(workbook, name)
    reference = validate_table_range(
        worksheet,
        name,
        spec.get("range", table.ref),
    )
    new_name = validate_table_name(spec.get("new_name", name), "update_table.new_name")
    if new_name != name and new_name in table_names(workbook):
        fail("ambiguous_edit", "Updated table name already exists.", table=new_name)
    style = (
        build_table_style(require_mapping(spec["style"], "update_table.style"))
        if "style" in spec
        else None
    )
    table.ref = reference
    if new_name != name:
        table.displayName = new_name
        table.name = new_name
    if style is not None:
        table.tableStyleInfo = style
    if "totals_row_shown" in spec:
        table.totalsRowShown = bool(spec["totals_row_shown"])


def edit_operation(
    workbook: Any,
    operation: Mapping[str, Any],
    *,
    base_directory: Path,
    budget: CellBudget,
) -> tuple[Counter[str], list[str]]:
    op = require_string(operation.get("op"), "operation.op")
    allowed_by_operation = {
        "set_cell": {"op", "sheet", "cell", "spec", "formula", "value", "allow_formula"},
        "set_range": {"op", "sheet", "range", "values"},
        "format_range": {"op", "sheet", "range", "style"},
        "add_sheet": {
            "op",
            "name",
            "index",
            "source",
            "rows",
            "cells",
            "state",
            "merges",
            "freeze_panes",
            "dimensions",
            "auto_filter",
            "auto_width",
            "tab_color",
            "show_gridlines",
            "zoom_scale",
            "tables",
            "conditional_formats",
            "data_validations",
            "charts",
        },
        "remove_sheet": {"op", "sheet", "allow_unupdated_dependencies"},
        "rename_sheet": {"op", "sheet", "new_name", "allow_unupdated_dependencies"},
        "set_sheet": {
            "op",
            "sheet",
            "state",
            "freeze_panes",
            "auto_filter",
            "auto_width",
            "dimensions",
            "merges",
            "unmerge",
            "allow_merge_data_loss",
            "tab_color",
            "show_gridlines",
            "zoom_scale",
        },
        "add_table": {"op", "sheet", "table"},
        "update_table": {"op", "table"},
        "add_chart": {"op", "sheet", "chart"},
        "update_chart": {"op", "sheet", "index", "chart"},
        "add_conditional_format": {"op", "sheet", "rule"},
        "add_data_validation": {"op", "sheet", "validation"},
        "set_properties": {"op", "properties"},
        "add_defined_name": {"op", "defined_name"},
        "insert_rows": {"op", "sheet", "index", "amount"},
        "delete_rows": {"op", "sheet", "index", "amount"},
        "move_rows": {"op", "sheet", "start", "amount", "target"},
        "insert_columns": {"op", "sheet", "index", "amount"},
        "delete_columns": {"op", "sheet", "index", "amount"},
        "move_columns": {"op", "sheet", "start", "amount", "target"},
        "move_range": {"op", "sheet", "range", "target"},
    }
    if op in allowed_by_operation:
        reject_unknown_keys(operation, allowed_by_operation[op], f"{op} operation")
    counts: Counter[str] = Counter()
    warnings: list[str] = []
    if op in {
        "insert_rows",
        "delete_rows",
        "move_rows",
        "insert_columns",
        "delete_columns",
        "move_columns",
        "move_range",
    }:
        fail(
            "ambiguous_edit",
            "Row, column, and range moves are refused because dependent formulas, tables, "
            "charts, and names cannot be updated reliably.",
            operation=op,
        )
    if op == "set_cell":
        worksheet = worksheet_by_name(
            workbook,
            require_string(operation.get("sheet"), "set_cell.sheet"),
        )
        coordinate = require_string(operation.get("cell"), "set_cell.cell")
        target = worksheet[coordinate]
        if not hasattr(target, "coordinate"):
            fail("bad_input", "set_cell.cell must identify one cell.", cell=coordinate)
        value_fields = [key for key in ("spec", "formula", "value") if key in operation]
        if len(value_fields) != 1:
            fail("bad_input", "set_cell requires exactly one of spec, formula, or value.")
        budget.consume(1, label="set_cell")
        if "spec" in operation:
            set_cell_from_spec(target, operation["spec"])
        elif "formula" in operation:
            formula_cell_value(target, operation["formula"])
        elif "value" in operation:
            literal_cell_value(
                target,
                operation["value"],
                allow_formula=bool(operation.get("allow_formula", False)),
            )
        counts["cells_set"] += 1
    elif op == "set_range":
        worksheet = worksheet_by_name(
            workbook,
            require_string(operation.get("sheet"), "set_range.sheet"),
        )
        counts["cells_set"] += set_range_values(
            worksheet,
            require_string(operation.get("range"), "set_range.range"),
            operation.get("values"),
            budget=budget,
        )
    elif op == "format_range":
        worksheet = worksheet_by_name(
            workbook,
            require_string(operation.get("sheet"), "format_range.sheet"),
        )
        counts["cells_formatted"] += format_range(
            worksheet,
            require_string(operation.get("range"), "format_range.range"),
            require_mapping(operation.get("style"), "format_range.style"),
            budget=budget,
        )
    elif op == "add_sheet":
        name = unique_sheet_title(
            workbook,
            require_string(operation.get("name"), "add_sheet.name"),
        )
        worksheet = workbook.create_sheet(name, operation.get("index"))
        sheet_counts, configure_warnings = configure_worksheet(
            workbook,
            worksheet,
            operation,
            base_directory=base_directory,
            budget=budget,
        )
        counts.update(sheet_counts)
        warnings.extend(configure_warnings)
        counts["sheets_added"] += 1
    elif op == "remove_sheet":
        name = require_string(operation.get("sheet"), "remove_sheet.sheet")
        worksheet = worksheet_by_name(workbook, name)
        warnings.extend(
            require_safe_sheet_change(
                workbook,
                name,
                acknowledged=bool(operation.get("allow_unupdated_dependencies", False)),
                operation="Sheet removal",
            )
        )
        visible = [sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"]
        if worksheet.sheet_state == "visible" and len(visible) == 1:
            fail("bad_input", "Cannot remove the workbook's last visible sheet.")
        workbook.remove(worksheet)
        counts["sheets_removed"] += 1
    elif op == "rename_sheet":
        old_name = require_string(operation.get("sheet"), "rename_sheet.sheet")
        worksheet = worksheet_by_name(workbook, old_name)
        new_name = unique_sheet_title(
            workbook,
            require_string(operation.get("new_name"), "rename_sheet.new_name"),
        )
        warnings.extend(
            require_safe_sheet_change(
                workbook,
                old_name,
                acknowledged=bool(operation.get("allow_unupdated_dependencies", False)),
                operation="Sheet rename",
            )
        )
        worksheet.title = new_name
        counts["sheets_renamed"] += 1
    elif op == "set_sheet":
        worksheet = worksheet_by_name(
            workbook,
            require_string(operation.get("sheet"), "set_sheet.sheet"),
        )
        if "state" in operation:
            state = operation["state"]
            if state not in {"visible", "hidden", "veryHidden"}:
                fail("bad_input", "Invalid worksheet state.", state=state)
            if state != "visible":
                visible = [
                    sheet
                    for sheet in workbook.worksheets
                    if sheet.sheet_state == "visible" and sheet is not worksheet
                ]
                if not visible:
                    fail("bad_input", "At least one worksheet must remain visible.")
            worksheet.sheet_state = state
        counts.update(
            apply_sheet_settings(
                worksheet,
                operation,
                context="edit",
                warnings=warnings,
            )
        )
        counts["sheets_updated"] += 1
    elif op == "add_table":
        worksheet = worksheet_by_name(
            workbook,
            require_string(operation.get("sheet"), "add_table.sheet"),
        )
        add_table(
            worksheet,
            require_mapping(operation.get("table"), "add_table.table"),
            workbook,
        )
        counts["tables_added"] += 1
    elif op == "update_table":
        update_table(
            workbook,
            require_mapping(operation.get("table"), "update_table.table"),
        )
        counts["tables_updated"] += 1
    elif op == "add_chart":
        worksheet = worksheet_by_name(
            workbook,
            require_string(operation.get("sheet"), "add_chart.sheet"),
        )
        add_chart(
            worksheet,
            require_mapping(operation.get("chart"), "add_chart.chart"),
            workbook,
        )
        counts["charts_added"] += 1
    elif op == "update_chart":
        worksheet = worksheet_by_name(
            workbook,
            require_string(operation.get("sheet"), "update_chart.sheet"),
        )
        index = operation.get("index")
        if not isinstance(index, int) or not 0 <= index < len(worksheet._charts):
            fail(
                "bad_input",
                "update_chart.index does not identify an existing chart.",
                index=index,
            )
        old_chart = worksheet._charts[index]
        chart_spec = dict(require_mapping(operation.get("chart"), "update_chart.chart"))
        chart_spec.setdefault("anchor", chart_anchor(old_chart) or "E2")
        add_chart(worksheet, chart_spec, workbook)
        replacement = worksheet._charts.pop()
        worksheet._charts[index] = replacement
        counts["charts_updated"] += 1
    elif op == "add_conditional_format":
        worksheet = worksheet_by_name(
            workbook,
            require_string(operation.get("sheet"), "conditional format sheet"),
        )
        add_conditional_format(
            worksheet,
            require_mapping(operation.get("rule"), "conditional format rule"),
        )
        counts["conditional_formats_added"] += 1
    elif op == "add_data_validation":
        worksheet = worksheet_by_name(
            workbook,
            require_string(operation.get("sheet"), "data validation sheet"),
        )
        add_validation(
            worksheet,
            require_mapping(operation.get("validation"), "data validation"),
        )
        counts["data_validations_added"] += 1
    elif op == "set_properties":
        set_workbook_properties(
            workbook,
            require_mapping(operation.get("properties"), "set_properties.properties"),
        )
        counts["properties_updated"] += 1
    elif op == "add_defined_name":
        add_defined_names(
            workbook,
            [require_mapping(operation.get("defined_name"), "defined name")],
        )
        counts["defined_names_added"] += 1
    else:
        fail("unsupported_operation", "Unsupported edit operation.", operation=op)
    return counts, warnings


def handle_edit(args: argparse.Namespace) -> dict[str, Any]:
    package = preflight_xlsx(args.source)
    job = load_json_object(args.job, "edit job")
    validate_job(job)
    reject_unknown_keys(job, {"schema_version", "operations"}, "edit job")
    workbook = open_workbook(args.source, data_only=False)
    operations = require_list(job.get("operations"), "operations")
    if not operations:
        fail("bad_input", "Edit job requires at least one operation.")
    total_counts: Counter[str] = Counter()
    budget = cell_budget_from_args(args)
    warnings: list[str] = mutation_preservation_warnings(workbook, "Edit")
    for index, raw_operation in enumerate(operations):
        operation = require_mapping(raw_operation, f"operations[{index}]")
        counts, operation_warnings = edit_operation(
            workbook,
            operation,
            base_directory=args.job.resolve().parent,
            budget=budget,
        )
        total_counts.update(counts)
        warnings.extend(operation_warnings)
    formula_count = sum(
        cell.data_type == "f"
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
    )
    if formula_count:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        warnings.append(
            "The result contains formulas; cached values may be stale and recalculation is only "
            "requested."
        )
    verification = save_workbook_atomic(
        workbook,
        args.output,
        source=args.source,
        overwrite=args.overwrite,
    )
    workbook.close()
    return success(
        "edit",
        {
            "source": str(args.source.resolve()),
            "job": str(args.job.resolve()),
            "output": str(args.output.resolve()),
            "operations": len(operations),
            "counts": dict(total_counts),
            "budgeted_cells": budget.used,
            "source_package": package,
            "verification": verification,
        },
        warnings,
    )


def dataframe_from_worksheet(
    worksheet: Any,
    *,
    range_text: str | None,
    header_row: int | None,
) -> tuple[Any, tuple[int, int, int, int]]:
    bounds = normalize_range(worksheet, range_text)
    min_row, min_col, max_row, max_col = bounds
    actual_header = header_row if header_row is not None else min_row
    if not min_row <= actual_header <= max_row:
        fail(
            "bad_input",
            "header_row must fall inside the selected range.",
            header_row=actual_header,
            range=range_to_a1(bounds),
        )
    headers = [
        worksheet.cell(actual_header, column).value for column in range(min_col, max_col + 1)
    ]
    if any(value is None or str(value).strip() == "" for value in headers):
        fail("bad_input", "Rectangular transformation requires non-empty column headers.")
    header_text = [str(value) for value in headers]
    if len(set(header_text)) != len(header_text):
        fail("bad_input", "Rectangular transformation requires unique column headers.")
    rows = [
        [worksheet.cell(row, column).value for column in range(min_col, max_col + 1)]
        for row in range(actual_header + 1, max_row + 1)
    ]
    return pd.DataFrame(rows, columns=header_text), bounds


def validate_columns(frame: Any, columns: Sequence[Any], label: str) -> list[str]:
    names = [require_string(column, label) for column in columns]
    missing = [name for name in names if name not in frame.columns]
    if missing:
        fail("bad_input", f"{label} references unknown columns.", columns=missing)
    return names


def clean_dataframe(frame: Any, policies: Mapping[str, Any]) -> tuple[Any, dict[str, int]]:
    reject_unknown_keys(
        policies,
        {
            "whitespace",
            "missing",
            "malformed_rows",
            "coercion",
            "rename",
            "duplicates",
            "sort",
            "na_position",
        },
        "policies",
    )
    result = frame.copy()
    counts: Counter[str] = Counter(input_rows=len(result))
    whitespace = require_mapping(policies.get("whitespace", {}), "policies.whitespace")
    if whitespace:
        reject_unknown_keys(
            whitespace,
            {"columns", "strip", "collapse"},
            "policies.whitespace",
        )
        columns = whitespace.get(
            "columns",
            [column for column in result.columns if result[column].dtype == object],
        )
        columns = validate_columns(result, columns, "whitespace.columns")
        strip = bool(whitespace.get("strip", True))
        collapse = bool(whitespace.get("collapse", False))
        for column in columns:

            def clean_text(value: Any) -> Any:
                if not isinstance(value, str):
                    return value
                cleaned = value.strip() if strip else value
                return re.sub(r"\s+", " ", cleaned) if collapse else cleaned

            before = result[column].copy()
            result[column] = result[column].map(clean_text)

            def equal_values(left: Any, right: Any) -> bool:
                left_missing = bool(pd.isna(left))
                right_missing = bool(pd.isna(right))
                if left_missing or right_missing:
                    return left_missing and right_missing
                return bool(left == right)

            counts["whitespace_values_changed"] += int(
                sum(
                    not equal_values(left, right)
                    for left, right in zip(before, result[column], strict=True)
                )
            )
    missing_policy = require_mapping(policies.get("missing", {}), "policies.missing")
    if missing_policy:
        reject_unknown_keys(
            missing_policy,
            {"markers", "fill", "required_columns", "drop_rows"},
            "policies.missing",
        )
        markers = require_list(missing_policy.get("markers", []), "missing.markers")
        if markers:
            result = result.replace(markers, pd.NA)
        fill = require_mapping(missing_policy.get("fill", {}), "missing.fill")
        for column, value in fill.items():
            validate_columns(result, [column], "missing.fill")
            result[column] = result[column].fillna(value)
        required = validate_columns(
            result,
            require_list(missing_policy.get("required_columns", []), "required_columns"),
            "missing.required_columns",
        )
        before_rows = len(result)
        if required:
            result = result.dropna(subset=required)
        drop_rows = missing_policy.get("drop_rows")
        if drop_rows is not None:
            if drop_rows not in {"any", "all"}:
                fail("bad_input", "missing.drop_rows must be 'any' or 'all'.")
            result = result.dropna(how=drop_rows)
        counts["missing_rows_dropped"] += before_rows - len(result)
    malformed = require_mapping(
        policies.get("malformed_rows", {}),
        "policies.malformed_rows",
    )
    if malformed:
        reject_unknown_keys(
            malformed,
            {"required_columns", "max_missing_fraction", "action"},
            "policies.malformed_rows",
        )
        required = validate_columns(
            result,
            require_list(malformed.get("required_columns", []), "required_columns"),
            "malformed_rows.required_columns",
        )
        before_rows = len(result)
        if required:
            invalid = result[required].isna().any(axis=1)
            action = malformed.get("action", "error")
            if invalid.any() and action == "error":
                fail(
                    "bad_input",
                    "Malformed rows violate required-column policy.",
                    row_indexes=[int(value) for value in result.index[invalid][:20]],
                )
            if action == "drop":
                result = result.loc[~invalid]
            elif action != "error":
                fail("bad_input", "malformed_rows.action must be 'error' or 'drop'.")
        max_missing_fraction = malformed.get("max_missing_fraction")
        if max_missing_fraction is not None:
            threshold = float(max_missing_fraction)
            if not 0 <= threshold <= 1:
                fail("bad_input", "max_missing_fraction must be between 0 and 1.")
            invalid = result.isna().mean(axis=1) > threshold
            action = malformed.get("action", "error")
            if invalid.any() and action == "error":
                fail(
                    "bad_input",
                    "Malformed rows exceed the missing-value threshold.",
                    row_indexes=[int(value) for value in result.index[invalid][:20]],
                )
            if action == "drop":
                result = result.loc[~invalid]
        counts["malformed_rows_dropped"] += before_rows - len(result)
    coercion = require_mapping(policies.get("coercion", {}), "policies.coercion")
    if coercion:
        result = apply_dataframe_schema(result, coercion)
        counts["columns_coerced"] += len(coercion)
    rename = require_mapping(policies.get("rename", {}), "policies.rename")
    if rename:
        validate_columns(result, list(rename), "rename")
        targets = [require_string(value, "rename target") for value in rename.values()]
        resulting_columns = [rename.get(column, column) for column in result.columns]
        if len(set(resulting_columns)) != len(resulting_columns):
            fail("bad_input", "Column renaming would create duplicate names.")
        result = result.rename(columns=rename)
        counts["columns_renamed"] += len(targets)
    duplicates = require_mapping(policies.get("duplicates", {}), "policies.duplicates")
    if duplicates:
        reject_unknown_keys(duplicates, {"subset", "keep"}, "policies.duplicates")
        subset_raw = duplicates.get("subset")
        subset = (
            validate_columns(
                result,
                require_list(subset_raw, "duplicates.subset"),
                "duplicates.subset",
            )
            if subset_raw is not None
            else None
        )
        keep = duplicates.get("keep", "first")
        if keep == "none":
            keep = False
        if keep not in {"first", "last", False}:
            fail("bad_input", "duplicates.keep must be first, last, or none.")
        before_rows = len(result)
        result = result.drop_duplicates(subset=subset, keep=keep)
        counts["duplicate_rows_dropped"] += before_rows - len(result)
    sort_specs = require_list(policies.get("sort", []), "policies.sort")
    if sort_specs:
        columns: list[str] = []
        ascending: list[bool] = []
        for index, raw_spec in enumerate(sort_specs):
            if isinstance(raw_spec, str):
                column, direction = raw_spec, True
            else:
                spec = require_mapping(raw_spec, f"sort[{index}]")
                reject_unknown_keys(spec, {"column", "ascending"}, f"sort[{index}]")
                column = require_string(spec.get("column"), "sort column")
                direction = bool(spec.get("ascending", True))
            validate_columns(result, [column], "sort")
            columns.append(column)
            ascending.append(direction)
        result = result.sort_values(
            by=columns,
            ascending=ascending,
            kind="stable",
            na_position=policies.get("na_position", "last"),
        )
        counts["sort_keys"] += len(columns)
    result = result.reset_index(drop=True)
    counts["output_rows"] = len(result)
    counts["output_columns"] = len(result.columns)
    return result, dict(counts)


def prepare_destination_sheet(
    workbook: Any,
    name: str,
    *,
    replace: bool,
    source_sheet: str,
) -> Any:
    if name == source_sheet:
        fail(
            "bad_input",
            "Rectangular results must use a separate sheet so source data remains preserved.",
            sheet=name,
        )
    if name in workbook.sheetnames:
        if not replace:
            fail(
                "ambiguous_edit",
                "Destination sheet already exists.",
                sheet=name,
                resolution="Set replace_destination to true or choose another name.",
            )
        target = workbook[name]
        if target.sheet_state == "visible":
            visible = [
                sheet
                for sheet in workbook.worksheets
                if sheet.sheet_state == "visible" and sheet is not target
            ]
            if not visible:
                fail("bad_input", "Cannot replace the workbook's only visible sheet.")
        index = workbook.worksheets.index(target)
        workbook.remove(target)
        return workbook.create_sheet(name, index)
    return workbook.create_sheet(unique_sheet_title(workbook, name))


def write_dataframe_sheet(
    worksheet: Any,
    frame: Any,
    *,
    start_row: int = 1,
    add_filter: bool = True,
) -> dict[str, int]:
    rows = dataframe_rows(frame)
    count = populate_rows(worksheet, rows, start_row=start_row)
    if frame.shape[1]:
        for cell in worksheet[start_row]:
            if cell.column <= frame.shape[1]:
                cell.font = Font(bold=True, color="FFFFFFFF")
                cell.fill = PatternFill("solid", fgColor="FF1F4E78")
                cell.alignment = Alignment(horizontal="center")
        for index, column in enumerate(frame.columns, start=1):
            sample_lengths = [len(str(column))]
            sample_lengths.extend(
                len(str(value))
                for value in frame[column].head(AUTO_WIDTH_DEFAULTS["sample_rows"])
                if not pd.isna(value)
            )
            # min_width 1 keeps the historical header-length-driven widths of result sheets.
            worksheet.column_dimensions[get_column_letter(index)].width = auto_width_value(
                sample_lengths,
                min_width=1,
                max_width=AUTO_WIDTH_DEFAULTS["max_width"],
            )
        end_row = start_row + len(frame)
        end_col = get_column_letter(frame.shape[1])
        if add_filter:
            worksheet.auto_filter.ref = f"A{start_row}:{end_col}{end_row}"
        worksheet.freeze_panes = f"A{start_row + 1}"
    return {"cells": count, "rows": len(frame), "columns": len(frame.columns)}


def handle_clean(args: argparse.Namespace) -> dict[str, Any]:
    package = preflight_xlsx(args.source)
    job = load_json_object(args.job, "clean job")
    validate_job(job)
    reject_unknown_keys(
        job,
        {
            "schema_version",
            "source",
            "destination_sheet",
            "replace_destination",
            "policies",
            "table",
        },
        "clean job",
    )
    source_spec = require_mapping(job.get("source"), "source")
    reject_unknown_keys(source_spec, {"sheet", "range", "header_row"}, "clean source")
    source_name = source_spec.get("sheet")
    formulas_workbook = open_workbook(args.source, data_only=False)
    values_workbook = open_workbook(args.source, data_only=True)
    formula_sheet = worksheet_by_name(formulas_workbook, source_name)
    values_sheet = worksheet_by_name(values_workbook, formula_sheet.title)
    frame, bounds = dataframe_from_worksheet(
        values_sheet,
        range_text=source_spec.get("range"),
        header_row=source_spec.get("header_row"),
    )
    policies = require_mapping(job.get("policies", {}), "policies")
    cleaned, counts = clean_dataframe(frame, policies)
    destination_name = require_string(job.get("destination_sheet"), "destination_sheet")
    destination = prepare_destination_sheet(
        formulas_workbook,
        destination_name,
        replace=bool(job.get("replace_destination", False)),
        source_sheet=formula_sheet.title,
    )
    written = write_dataframe_sheet(destination, cleaned)
    if job.get("table"):
        table_spec = dict(require_mapping(job["table"], "table"))
        table_spec.setdefault("name", re.sub(r"\W+", "_", destination_name) + "_Clean")
        table_spec.setdefault(
            "range",
            f"A1:{get_column_letter(len(cleaned.columns))}{len(cleaned) + 1}",
        )
        add_table(destination, table_spec, formulas_workbook)
    formula_cells = sum(
        cell.data_type == "f"
        for row in formula_sheet.iter_rows(
            min_row=bounds[0],
            min_col=bounds[1],
            max_row=bounds[2],
            max_col=bounds[3],
        )
        for cell in row
    )
    warnings = mutation_preservation_warnings(formulas_workbook, "Clean")
    warnings.append(
        "Cleanup is interpretive; the unchanged source data remains on "
        f"sheet {formula_sheet.title!r}."
    )
    if formula_cells:
        warnings.append(
            "Cleanup used cached values for source formulas; those values may be stale or absent."
        )
    verification = save_workbook_atomic(
        formulas_workbook,
        args.output,
        source=args.source,
        overwrite=args.overwrite,
    )
    formulas_workbook.close()
    values_workbook.close()
    return success(
        "clean",
        {
            "source": str(args.source.resolve()),
            "job": str(args.job.resolve()),
            "output": str(args.output.resolve()),
            "source_sheet": formula_sheet.title,
            "source_range": range_to_a1(bounds),
            "destination_sheet": destination_name,
            "counts": {**counts, "cells_written": written["cells"]},
            "source_package": package,
            "verification": verification,
        },
        warnings,
    )


def aggregate_function(name: str) -> str | Callable[[Any], Any]:
    supported = {
        "sum": "sum",
        "mean": "mean",
        "median": "median",
        "min": "min",
        "max": "max",
        "count": "count",
        "nunique": "nunique",
        "first": "first",
        "last": "last",
    }
    if name not in supported:
        fail("bad_input", "Unsupported summary aggregation.", aggregation=name)
    return supported[name]


def build_static_summary(frame: Any, job: Mapping[str, Any]) -> Any:
    group_by = validate_columns(
        frame,
        require_list(job.get("group_by"), "group_by"),
        "group_by",
    )
    if not group_by:
        fail("bad_input", "summarize requires at least one group_by column.")
    columns = validate_columns(
        frame,
        require_list(job.get("columns", []), "columns"),
        "columns",
    )
    raw_values = require_list(job.get("values"), "values")
    if not raw_values:
        fail("bad_input", "summarize requires at least one value aggregation.")
    aggregations: dict[str, list[str]] = {}
    requested_names: dict[tuple[str, str], str] = {}
    for index, raw_value in enumerate(raw_values):
        spec = require_mapping(raw_value, f"values[{index}]")
        reject_unknown_keys(
            spec,
            {"column", "aggregation", "name"},
            f"values[{index}]",
        )
        column = require_string(spec.get("column"), "summary value column")
        validate_columns(frame, [column], "values")
        aggregation = require_string(spec.get("aggregation", "sum"), "aggregation")
        aggregate_function(aggregation)
        aggregations.setdefault(column, []).append(aggregation)
        if "name" in spec:
            requested_names[(column, aggregation)] = require_string(spec["name"], "name")
    try:
        summary = pd.pivot_table(
            frame,
            index=group_by,
            columns=columns or None,
            values=list(aggregations),
            aggfunc=aggregations,
            fill_value=job.get("fill_value"),
            dropna=bool(job.get("dropna", True)),
            sort=bool(job.get("sort", True)),
            observed=bool(job.get("observed", True)),
        )
    except (KeyError, TypeError, ValueError) as exc:
        fail("bad_input", "Could not create static summary.", reason=str(exc))
    summary = summary.reset_index()
    flattened: list[str] = []
    for raw_column in summary.columns:
        column_parts = raw_column if isinstance(raw_column, tuple) else (raw_column,)
        parts = [str(part) for part in column_parts]
        parts = [part for part in parts if part and part != "None"]
        if len(parts) >= 2 and (parts[0], parts[1]) in requested_names:
            base = requested_names[(parts[0], parts[1])]
            parts = [base, *parts[2:]]
        flattened.append(" | ".join(parts))
    if len(set(flattened)) != len(flattened):
        fail(
            "ambiguous_edit",
            "Summary column labels are not unique; provide distinct value names.",
            columns=flattened,
        )
    summary.columns = flattened
    return summary


def handle_summarize(args: argparse.Namespace) -> dict[str, Any]:
    package = preflight_xlsx(args.source)
    job = load_json_object(args.job, "summarize job")
    validate_job(job)
    reject_unknown_keys(
        job,
        {
            "schema_version",
            "source",
            "destination_sheet",
            "replace_destination",
            "group_by",
            "columns",
            "values",
            "fill_value",
            "dropna",
            "sort",
            "observed",
            "chart",
        },
        "summarize job",
    )
    source_spec = require_mapping(job.get("source"), "source")
    reject_unknown_keys(source_spec, {"sheet", "range", "header_row"}, "summarize source")
    formulas_workbook = open_workbook(args.source, data_only=False)
    values_workbook = open_workbook(args.source, data_only=True)
    formula_sheet = worksheet_by_name(formulas_workbook, source_spec.get("sheet"))
    values_sheet = worksheet_by_name(values_workbook, formula_sheet.title)
    frame, bounds = dataframe_from_worksheet(
        values_sheet,
        range_text=source_spec.get("range"),
        header_row=source_spec.get("header_row"),
    )
    summary = build_static_summary(frame, job)
    destination_name = require_string(job.get("destination_sheet"), "destination_sheet")
    destination = prepare_destination_sheet(
        formulas_workbook,
        destination_name,
        replace=bool(job.get("replace_destination", False)),
        source_sheet=formula_sheet.title,
    )
    destination["A1"] = "STATIC SUMMARY — values do not refresh automatically"
    destination["A1"].font = Font(bold=True, color="FFC00000")
    destination.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=max(1, len(summary.columns)),
    )
    written = write_dataframe_sheet(destination, summary, start_row=3)
    chart_count = 0
    if "chart" in job:
        chart_spec = dict(require_mapping(job["chart"], "chart"))
        if len(summary.columns) < 2 or len(summary) < 1:
            fail("bad_input", "A summary chart requires at least two columns and one row.")
        chart_spec.setdefault("type", "bar")
        chart_spec.setdefault(
            "data",
            f"B3:{get_column_letter(len(summary.columns))}{len(summary) + 3}",
        )
        chart_spec.setdefault("categories", f"A4:A{len(summary) + 3}")
        chart_spec.setdefault("anchor", f"A{len(summary) + 6}")
        add_chart(destination, chart_spec, formulas_workbook)
        chart_count = 1
    formula_cells = sum(
        cell.data_type == "f"
        for row in formula_sheet.iter_rows(
            min_row=bounds[0],
            min_col=bounds[1],
            max_row=bounds[2],
            max_col=bounds[3],
        )
        for cell in row
    )
    warnings = mutation_preservation_warnings(formulas_workbook, "Summarize")
    warnings.append(
        "The generated grouped/pivot-style report is static, not a native Excel pivot table."
    )
    if formula_cells:
        warnings.append(
            "Summary input used cached formula values; those values may be stale or absent."
        )
    verification = save_workbook_atomic(
        formulas_workbook,
        args.output,
        source=args.source,
        overwrite=args.overwrite,
    )
    formulas_workbook.close()
    values_workbook.close()
    return success(
        "summarize",
        {
            "source": str(args.source.resolve()),
            "job": str(args.job.resolve()),
            "output": str(args.output.resolve()),
            "source_sheet": formula_sheet.title,
            "source_range": range_to_a1(bounds),
            "destination_sheet": destination_name,
            "rows": len(summary),
            "columns": len(summary.columns),
            "cells_written": written["cells"],
            "charts": chart_count,
            "source_package": package,
            "verification": verification,
        },
        warnings,
    )


def infer_format(path: Path, explicit: str, *, allowed: set[str]) -> str:
    if explicit != "auto":
        return explicit
    inferred = path.suffix.lower().lstrip(".")
    if inferred not in allowed:
        fail(
            "bad_input",
            "Could not infer format from the path extension; select it explicitly.",
            path=path,
            allowed=sorted(allowed),
        )
    return inferred


def convert_frame_for_export(
    frame: Any,
    *,
    allow_formulas: bool,
) -> tuple[Any, int]:
    result = frame.copy()
    guarded = 0
    for index in range(result.shape[1]):
        output: list[Any] = []
        for value in result.iloc[:, index]:
            sanitized, changed = sanitize_formula_text(
                value,
                allow_formulas=allow_formulas,
            )
            output.append(sanitized)
            guarded += changed
        result.isetitem(index, output)
    output_columns: list[Any] = []
    for column in result.columns:
        sanitized, changed = sanitize_formula_text(
            column,
            allow_formulas=allow_formulas,
        )
        output_columns.append(sanitized)
        guarded += changed
    result.columns = output_columns
    return result, guarded


def write_frame_delimited(
    frame: Any,
    handle: Any,
    *,
    delimiter: str,
    args: argparse.Namespace,
) -> None:
    frame.to_csv(
        handle,
        sep=delimiter,
        index=args.index,
        header=True,
        na_rep=args.na_value,
        date_format=args.date_format,
        quoting=csv_quoting(args.quoting),
        escapechar="\\" if args.quoting == "none" else None,
        lineterminator="\n",
    )


def export_one_sheet(
    workbook: Any,
    worksheet: Any,
    output: Path,
    *,
    source: Path,
    output_format: str,
    args: argparse.Namespace,
    overwrite: bool,
) -> tuple[dict[str, Any], int, int, int]:
    delimiter = args.delimiter or ("\t" if output_format == "tsv" else ",")
    if len(delimiter) != 1:
        fail("bad_input", "--delimiter must be exactly one character.")
    if args.sheet_policy == "raw":
        if args.header_row is not None or args.schema or args.index:
            fail(
                "bad_input",
                "Raw-sheet export cannot use --header-row, --schema, or --index; "
                "select --sheet-policy header.",
            )
        bounds = normalize_range(worksheet, args.range)
        is_blank = args.range is None and not any(
            cell.value is not None for cell in worksheet._cells.values()
        )
        rows = [] if is_blank else list(iter_range_values(worksheet, bounds))
        output_rows: list[list[Any]] = []
        guarded = 0
        for row in rows:
            output_row: list[Any] = []
            for value in row:
                if args.date_format and isinstance(value, (dt.datetime, dt.date)):
                    value = value.strftime(args.date_format)
                value, changed = sanitize_formula_text(
                    value,
                    allow_formulas=args.allow_formulas,
                )
                output_row.append(serialize_cell_value(value))
                guarded += changed
            output_rows.append(output_row)

        def writer(handle: Any) -> None:
            csv_writer = csv.writer(
                handle,
                delimiter=delimiter,
                quoting=csv_quoting(args.quoting),
                lineterminator="\n",
                escapechar="\\" if args.quoting == "none" else None,
            )
            csv_writer.writerows(
                [
                    [args.na_value if value is None else value for value in row]
                    for row in output_rows
                ]
            )

        row_count = len(output_rows)
        column_count = max((len(row) for row in output_rows), default=0)
    else:
        frame, _ = dataframe_from_worksheet(
            worksheet,
            range_text=args.range,
            header_row=args.header_row,
        )
        if args.schema:
            schema = load_json_object(args.schema, "schema file")
            frame = apply_dataframe_schema(frame, schema)
        frame, guarded = convert_frame_for_export(
            frame,
            allow_formulas=args.allow_formulas,
        )

        def writer(handle: Any) -> None:
            write_frame_delimited(frame, handle, delimiter=delimiter, args=args)

        row_count = len(frame) + 1
        column_count = len(frame.columns) + bool(args.index)

    verification = write_text_atomic(
        output,
        writer,
        source=source,
        overwrite=overwrite,
        encoding=args.encoding,
    )
    return verification, row_count, column_count, guarded


def safe_sheet_filename(index: int, sheet_name: str, extension: str) -> str:
    stem = re.sub(r"[^A-Za-z0-9._-]+", "_", sheet_name).strip("._") or "sheet"
    return f"{index:03d}-{stem}.{extension}"


def publish_directory_atomic(temp: Path, output: Path) -> None:
    if output.exists():
        fail(
            "unsupported_operation",
            "Atomic one-file-per-sheet publication requires a destination directory that does "
            "not already exist.",
            path=output,
        )
    try:
        os.replace(temp, output)
    except OSError as exc:
        shutil.rmtree(temp, ignore_errors=True)
        fail(
            "post_write_validation",
            "Could not atomically publish the output directory.",
            path=output,
            reason=str(exc),
        )


def require_external_timeout(value: Any) -> float:
    timeout = float(value)
    if not math.isfinite(timeout) or not 1 <= timeout <= 1800:
        fail("bad_input", "--timeout must be between 1 and 1800 seconds.", timeout=value)
    return timeout


def require_pypdfium2() -> Any:
    try:
        import pypdfium2
    except ModuleNotFoundError as exc:
        raise ToolError(
            "missing_dependency",
            "Page rendering requires the optional pypdfium2 package.",
            details={"install": f'"{sys.executable}" -m pip install pypdfium2'},
        ) from exc
    return pypdfium2


def _find_libreoffice() -> str:
    executable = shutil.which("soffice") or shutil.which("libreoffice")
    if executable is None:
        fail(
            "missing_dependency",
            "LibreOffice is required for PDF conversion and rendering but was not found on PATH.",
            executables=["soffice", "libreoffice"],
        )
    return executable


def _libreoffice_environment(work: Path) -> dict[str, str]:
    environment = {
        key: value
        for key, value in os.environ.items()
        if key in {"PATH", "LANG", "LC_ALL", "LC_CTYPE", "TZ", "SYSTEMROOT", "WINDIR"}
    }
    environment.update(
        {
            "HOME": str(work),
            "TMPDIR": str(work),
            "TMP": str(work),
            "TEMP": str(work),
            "XDG_CACHE_HOME": str(work / "cache"),
            "XDG_CONFIG_HOME": str(work / "config"),
        }
    )
    return environment


def _redact_diagnostic(text: str, paths: Sequence[Path]) -> str:
    redacted = text[-4000:]
    replacements: set[str] = set()
    for path in paths:
        try:
            resolved = path.resolve()
            replacements.update({str(path), str(resolved), resolved.as_uri(), path.name})
        except (OSError, ValueError):
            replacements.add(str(path))
    for value in sorted(replacements, key=len, reverse=True):
        if value:
            redacted = redacted.replace(value, "<redacted-path>")
    redacted = re.sub(
        r"(?i)\b[a-z][a-z0-9+.-]*://[^\s<>'\"]+",
        "<redacted-url>",
        redacted,
    )
    redacted = re.sub(
        r"(?i)\b(api[_-]?key|token|password|passwd|secret|authorization|session[_-]?id)"
        r"\s*[:=]\s*[^\s&;,]+",
        r"\1=<redacted>",
        redacted,
    )
    return redacted


@dataclass
class LibreOfficePdf:
    generated: Path
    stdout: str
    stderr: str
    office_version: str
    engine: str


def _run_libreoffice_pdf(
    source: Path,
    work: Path,
    timeout: float,
    redact_paths: Sequence[Path],
) -> LibreOfficePdf:
    """Convert one snapshot copy inside the work directory to PDF via LibreOffice."""

    executable = _find_libreoffice()
    output_dir = work / "output"
    profile = work / "profile"
    output_dir.mkdir()
    profile.mkdir()
    environment = _libreoffice_environment(work)
    command = [
        executable,
        "--headless",
        "--nologo",
        "--nodefault",
        "--nolockcheck",
        "--nofirststartwizard",
        f"-env:UserInstallation={profile.resolve().as_uri()}",
        "--convert-to",
        "pdf",
        "--outdir",
        str(output_dir),
        str(source),
    ]
    try:
        completed = subprocess.run(
            command,
            check=False,
            capture_output=True,
            text=True,
            timeout=timeout,
            env=environment,
        )
    except subprocess.TimeoutExpired as exc:
        raise ToolError(
            "external_tool_failure",
            "LibreOffice conversion timed out.",
            details={"timeout_seconds": timeout},
        ) from exc
    except OSError as exc:
        raise ToolError(
            "external_tool_failure",
            "LibreOffice could not be launched.",
            details={"reason": str(exc)},
        ) from exc
    diagnostic_paths = [source, work, output_dir, profile, *redact_paths]
    stdout = _redact_diagnostic(completed.stdout, diagnostic_paths)
    stderr = _redact_diagnostic(completed.stderr, diagnostic_paths)
    if completed.returncode != 0:
        raise ToolError(
            "external_tool_failure",
            "LibreOffice conversion failed.",
            details={"returncode": completed.returncode, "stdout": stdout, "stderr": stderr},
        )
    generated = output_dir / f"{source.stem}.pdf"
    if not generated.is_file():
        candidates = sorted(output_dir.glob("*.pdf"))
        if len(candidates) != 1:
            raise ToolError(
                "external_tool_failure",
                "LibreOffice did not produce exactly one PDF.",
                details={"outputs": [path.name for path in candidates]},
            )
        generated = candidates[0]
    try:
        version_result = subprocess.run(
            [executable, "--version"],
            check=False,
            capture_output=True,
            text=True,
            timeout=10,
            env=environment,
        )
        office_version = _redact_diagnostic(version_result.stdout.strip(), diagnostic_paths)
    except (OSError, subprocess.TimeoutExpired):
        office_version = "unknown"
    return LibreOfficePdf(
        generated=generated,
        stdout=stdout,
        stderr=stderr,
        office_version=office_version or "unknown",
        engine=Path(executable).name,
    )


def _validate_pdf_output(path: Path) -> int:
    try:
        with path.open("rb") as handle:
            if handle.read(5) != b"%PDF-":
                raise ToolError(
                    "external_tool_failure",
                    "LibreOffice output does not have a PDF signature.",
                )
        reader = pypdf.PdfReader(str(path), strict=True)
        pages = len(reader.pages)
    except ToolError:
        raise
    except Exception as exc:
        raise ToolError(
            "external_tool_failure",
            "LibreOffice PDF output could not be opened.",
            details={"reason": str(exc)},
        ) from exc
    if pages < 1:
        raise ToolError("post_write_validation", "Converted PDF contains no pages.")
    return pages


def workbook_pdf_context(source: Path) -> dict[str, Any]:
    """Inventory the workbook facts PDF conversion and rendering must report."""

    workbook = open_workbook(source, data_only=False)
    external_links = inventory_external_links(workbook)
    if external_links:
        fail(
            "unsupported_operation",
            "PDF conversion and rendering refuse workbooks with external links because "
            "LibreOffice may attempt to resolve them.",
            external_links=len(external_links),
        )
    context = {
        "sheet_names": list(workbook.sheetnames),
        "hidden_sheets": [
            worksheet.title
            for worksheet in workbook.worksheets
            if worksheet.sheet_state != "visible"
        ],
        "formula_cells": sum(
            cell.data_type == "f"
            for worksheet in workbook.worksheets
            for row in worksheet.iter_rows()
            for cell in row
        ),
        "fonts": workbook_font_inventory(workbook, source),
    }
    workbook.close()
    return context


def print_semantics_warnings(context: Mapping[str, Any], artifact: str) -> list[str]:
    warnings = font_portability_warnings(context["fonts"])
    if context["formula_cells"]:
        warnings.append(
            f"{artifact} values come from LibreOffice's load-time calculation of stored "
            "formulas; they are not proven to match another application's results."
        )
    if context["hidden_sheets"]:
        warnings.append(
            f"{len(context['hidden_sheets'])} hidden sheet(s) were excluded from the "
            f"{artifact}: " + ", ".join(context["hidden_sheets"]) + "."
        )
    warnings.append(
        f"{artifact} pagination follows LibreOffice print semantics (print areas, page "
        "setup, printed gridlines, fit settings); it is not proven identical to Excel's."
    )
    return warnings


def convert_xlsx_to_pdf(args: argparse.Namespace) -> dict[str, Any]:
    timeout = require_external_timeout(args.timeout)
    package = preflight_xlsx(args.source)
    source_hash = sha256_file(args.source)
    context = workbook_pdf_context(args.source)
    require_output_path(
        args.output,
        source=args.source,
        overwrite=args.overwrite,
        suffix=".pdf",
    )
    temporary = temporary_sibling(args.output, ".pdf")
    try:
        with tempfile.TemporaryDirectory(prefix="xlsx-libreoffice-") as temp_dir:
            work = Path(temp_dir)
            snapshot = work / args.source.name
            shutil.copyfile(args.source, snapshot)
            office = _run_libreoffice_pdf(
                snapshot,
                work,
                timeout,
                [args.source, args.output, temporary],
            )
            shutil.copyfile(office.generated, temporary)
        pages = _validate_pdf_output(temporary)
        if sha256_file(args.source) != source_hash:
            fail(
                "post_write_validation",
                "Source changed before the destination publication gate.",
                path=args.source,
            )
        atomic_publish_file(temporary, args.output)
    finally:
        temporary.unlink(missing_ok=True)
    warnings = print_semantics_warnings(context, "PDF")
    if office.stderr.strip():
        warnings.append("LibreOffice emitted diagnostics; inspect conversion.diagnostics.")
    return success(
        "convert",
        {
            "source": str(args.source.resolve()),
            "output": str(args.output.resolve()),
            "input_format": "xlsx",
            "output_format": "pdf",
            "counts": {
                "sheets": len(context["sheet_names"]),
                "hidden_sheets": len(context["hidden_sheets"]),
                "pdf_pages": pages,
            },
            "fonts": context["fonts"],
            "conversion": {
                "engine": office.engine,
                "libreoffice": office.office_version,
                "timeout_seconds": timeout,
                "diagnostics": {"stdout": office.stdout, "stderr": office.stderr},
            },
            "source_package": package,
            "verification": {
                "atomic_publish": True,
                "pdf_signature": True,
                "pdf_openable": True,
                "pdf_pages": pages,
                "source_unchanged_at_publish_gate": True,
            },
        },
        warnings,
    )


def convert_xlsx_to_delimited(
    args: argparse.Namespace,
    output_format: str,
) -> dict[str, Any]:
    package = preflight_xlsx(args.source)
    workbook = open_workbook(args.source, data_only=args.values == "cached")
    warnings: list[str] = []
    if args.values == "cached":
        warnings.append(
            "Delimited output uses cached formula values, which may be stale or absent."
        )
    if args.all_sheets:
        if args.sheet or args.range:
            fail("bad_input", "--all-sheets cannot be combined with --sheet or --range.")
        if args.output.suffix:
            fail("bad_input", "--all-sheets output must be a directory path without a suffix.")
        require_output_path(
            args.output,
            source=args.source,
            overwrite=args.overwrite,
        )
        if args.output.exists():
            fail(
                "unsupported_operation",
                "For atomic all-sheet export, choose a new destination directory.",
                path=args.output,
            )
        temp_directory = Path(
            tempfile.mkdtemp(prefix=f".{args.output.name}.", dir=args.output.parent)
        )
        outputs: list[dict[str, Any]] = []
        guarded_total = 0
        try:
            for index, worksheet in enumerate(workbook.worksheets, start=1):
                output = temp_directory / safe_sheet_filename(
                    index,
                    worksheet.title,
                    output_format,
                )
                verification, rows, columns, guarded = export_one_sheet(
                    workbook,
                    worksheet,
                    output,
                    source=args.source,
                    output_format=output_format,
                    args=args,
                    overwrite=False,
                )
                outputs.append(
                    {
                        "sheet": worksheet.title,
                        "file": output.name,
                        "sheet_policy": args.sheet_policy,
                        "rows": rows,
                        "columns": columns,
                        "verification": verification,
                    }
                )
                guarded_total += guarded
            publish_directory_atomic(temp_directory, args.output)
        except Exception:
            shutil.rmtree(temp_directory, ignore_errors=True)
            raise
        if guarded_total:
            warnings.append(
                f"Prefixed {guarded_total} potentially executable text value(s) with an apostrophe."
            )
        workbook.close()
        return success(
            "convert",
            {
                "source": str(args.source.resolve()),
                "output": str(args.output.resolve()),
                "input_format": "xlsx",
                "output_format": output_format,
                "all_sheets": True,
                "sheet_policy": args.sheet_policy,
                "files": outputs,
                "formula_injection_guarded": guarded_total,
                "source_package": package,
                "verification": {
                    "published_directory": args.output.is_dir(),
                    "file_count": len(outputs),
                },
            },
            warnings,
        )
    worksheet = worksheet_by_name(workbook, args.sheet)
    expected_suffix = f".{output_format}"
    if args.output.suffix.lower() != expected_suffix:
        fail("bad_input", f"Output path must use the {expected_suffix} extension.")
    verification, rows, columns, guarded = export_one_sheet(
        workbook,
        worksheet,
        args.output,
        source=args.source,
        output_format=output_format,
        args=args,
        overwrite=args.overwrite,
    )
    if guarded:
        warnings.append(
            f"Prefixed {guarded} potentially executable text value(s) with an apostrophe."
        )
    workbook.close()
    return success(
        "convert",
        {
            "source": str(args.source.resolve()),
            "output": str(args.output.resolve()),
            "input_format": "xlsx",
            "output_format": output_format,
            "sheet": worksheet.title,
            "sheet_policy": args.sheet_policy,
            "rows": rows,
            "columns": columns,
            "formula_injection_guarded": guarded,
            "source_package": package,
            "verification": verification,
        },
        warnings,
    )


def convert_delimited_to_xlsx(
    args: argparse.Namespace,
    input_format: str,
) -> dict[str, Any]:
    if args.output.suffix.lower() != ".xlsx":
        fail("bad_input", "Delimited conversion output must use the .xlsx extension.")
    if args.input_na_policy == "custom" and not args.input_na_value:
        fail(
            "bad_input",
            "--input-na-policy custom requires at least one --input-na-value.",
        )
    if args.input_na_policy != "custom" and args.input_na_value:
        fail(
            "bad_input",
            "--input-na-value requires --input-na-policy custom.",
        )
    source_spec: dict[str, Any] = {
        "path": str(args.source),
        "format": input_format,
        "encoding": args.encoding,
        "delimiter": args.delimiter or ("\t" if input_format == "tsv" else ","),
        "header": not args.no_header,
        "leading_zeros": args.leading_zeros,
        "malformed_rows": args.malformed_rows,
        "quoting": args.quoting,
        "recognize_default_na": args.input_na_policy == "default",
    }
    if args.input_na_policy == "custom":
        source_spec["na_values"] = args.input_na_value
    if args.schema:
        source_spec["schema"] = load_json_object(args.schema, "schema file")
    frame, source_inventory = read_delimited_frame(
        source_spec,
        base_directory=Path.cwd(),
    )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = args.sheet or "Data"
    guarded = 0
    rows = dataframe_rows(frame, include_header=not args.no_header)
    for row_index, row in enumerate(rows, start=1):
        for column_index, value in enumerate(row, start=1):
            cell = worksheet.cell(row_index, column_index)
            dangerous = is_dangerous_formula_text(value)
            if (
                args.allow_formulas
                and (args.no_header or row_index > 1)
                and isinstance(value, str)
                and value.startswith("=")
            ):
                formula_cell_value(cell, value)
            else:
                literal_cell_value(cell, value)
                guarded += bool(dangerous and not args.allow_formulas)
    if len(frame.columns) and not args.no_header:
        for cell in worksheet[1]:
            if cell.column <= len(frame.columns):
                cell.font = Font(bold=True, color="FFFFFFFF")
                cell.fill = PatternFill("solid", fgColor="FF1F4E78")
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(frame.columns))}{len(frame) + 1}"
    if any(cell.data_type == "f" for row in worksheet.iter_rows() for cell in row):
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
    verification = save_workbook_atomic(
        workbook,
        args.output,
        source=args.source,
        overwrite=args.overwrite,
    )
    workbook.close()
    warnings: list[str] = []
    if guarded:
        warnings.append(
            f"Stored {guarded} potentially executable delimited value(s) as literal text."
        )
    if args.allow_formulas:
        warnings.append(
            "Executable formulas were explicitly enabled for values beginning with '='; "
            "openpyxl did not calculate them."
        )
    return success(
        "convert",
        {
            "source": str(args.source.resolve()),
            "output": str(args.output.resolve()),
            "input_format": input_format,
            "output_format": "xlsx",
            "sheet": worksheet.title,
            "rows": len(frame),
            "columns": len(frame.columns),
            "formula_injection_guarded": guarded,
            "header_policy": "none" if args.no_header else "first_row",
            "input_na_policy": args.input_na_policy,
            "input_na_values": args.input_na_value,
            "source_inventory": source_inventory,
            "verification": verification,
        },
        warnings,
    )


def handle_convert(args: argparse.Namespace) -> dict[str, Any]:
    input_format = infer_format(
        args.source,
        args.input_format,
        allowed={"xlsx", "csv", "tsv"},
    )
    output_allowed = {"xlsx"} if input_format in {"csv", "tsv"} else {"csv", "tsv", "pdf"}
    output_format = (
        infer_format(
            args.output,
            args.output_format,
            allowed=output_allowed,
        )
        if not args.all_sheets
        else (args.output_format if args.output_format != "auto" else "csv")
    )
    if output_format != "pdf" and args.timeout != DEFAULT_EXTERNAL_TOOL_TIMEOUT:
        fail("bad_input", "--timeout applies only to PDF output.")
    if input_format == "xlsx" and output_format == "pdf":
        if (
            args.sheet is not None
            or args.range is not None
            or args.header_row is not None
            or args.sheet_policy is not None
            or args.all_sheets
            or args.values != "cached"
            or args.schema is not None
            or args.leading_zeros is not None
            or args.index
            or args.no_header
            or args.input_na_policy != "literal"
            or args.input_na_value
            or args.malformed_rows != "error"
            or args.encoding != "utf-8"
            or args.delimiter is not None
            or args.quoting != "minimal"
            or args.na_value != ""
            or args.date_format is not None
            or args.allow_formulas
        ):
            fail(
                "bad_input",
                "PDF conversion accepts only --timeout and --overwrite; delimited and "
                "sheet-selection options do not apply.",
            )
        return convert_xlsx_to_pdf(args)
    if input_format == "xlsx" and output_format in {"csv", "tsv"}:
        if (
            args.no_header
            or args.input_na_policy != "literal"
            or args.input_na_value
            or args.schema is not None
            or args.leading_zeros is not None
            or args.malformed_rows != "error"
        ):
            fail(
                "bad_input",
                "--schema, --leading-zeros, --no-header, input NA options, and "
                "--malformed-rows apply only to CSV/TSV input.",
            )
        args.sheet_policy = args.sheet_policy or "raw"
        if args.sheet_policy == "raw" and args.header_row is not None:
            fail(
                "bad_input",
                "--header-row requires --sheet-policy header.",
            )
        return convert_xlsx_to_delimited(args, output_format)
    if input_format in {"csv", "tsv"} and output_format == "xlsx":
        if (
            args.range is not None
            or args.header_row is not None
            or args.sheet_policy is not None
            or args.all_sheets
            or args.values != "cached"
            or args.index
            or args.na_value
            or args.date_format is not None
        ):
            fail(
                "bad_input",
                "--range, --header-row, --sheet-policy, --all-sheets, --values, "
                "--index, --na-value, and --date-format apply only to XLSX input.",
            )
        args.leading_zeros = args.leading_zeros or "preserve"
        return convert_delimited_to_xlsx(args, input_format)
    fail(
        "unsupported_operation",
        "Unsupported conversion direction.",
        input_format=input_format,
        output_format=output_format,
    )


def parse_render_pages(value: str | None, page_count: int) -> list[int]:
    if value is None:
        selected = list(range(1, page_count + 1))
    else:
        chosen: set[int] = set()
        for part in value.split(","):
            text = part.strip()
            if not text.isdigit():
                fail(
                    "bad_input",
                    "--pages must be a comma-separated list of 1-based page numbers.",
                    pages=value,
                )
            page = int(text)
            if not 1 <= page <= page_count:
                fail(
                    "bad_input",
                    "Requested page does not exist in the converted PDF.",
                    page=page,
                    pdf_pages=page_count,
                )
            chosen.add(page)
        if not chosen:
            fail("bad_input", "--pages must select at least one page.")
        selected = sorted(chosen)
    if len(selected) > MAX_RENDER_PAGES:
        fail(
            "resource_limit",
            "Rendering would produce too many pages; select --pages, set print areas, or "
            "hide sheets.",
            pages=len(selected),
            limit=MAX_RENDER_PAGES,
        )
    return selected


def handle_render(args: argparse.Namespace) -> dict[str, Any]:
    dpi = args.dpi
    if (
        isinstance(dpi, bool)
        or not isinstance(dpi, int)
        or not MIN_RENDER_DPI <= dpi <= MAX_RENDER_DPI
    ):
        fail(
            "bad_input",
            f"--dpi must be an integer between {MIN_RENDER_DPI} and {MAX_RENDER_DPI}.",
            dpi=dpi,
        )
    timeout = require_external_timeout(args.timeout)
    pdfium = require_pypdfium2()
    package = preflight_xlsx(args.source)
    source_hash = sha256_file(args.source)
    context = workbook_pdf_context(args.source)
    output = args.output
    if output.suffix:
        fail(
            "bad_input",
            "Render output must be a directory path without a suffix.",
            path=output,
        )
    if output.exists():
        fail(
            "unsupported_operation",
            "Atomic render publication requires a destination directory that does not "
            "already exist.",
            path=output,
        )
    if not output.parent.exists() or not output.parent.is_dir():
        fail("bad_input", "Destination parent directory does not exist.", path=output.parent)
    images: list[dict[str, Any]] = []
    staged = Path(tempfile.mkdtemp(prefix=f".{output.name}.", dir=output.parent))
    try:
        with tempfile.TemporaryDirectory(prefix="xlsx-libreoffice-") as temp_dir:
            work = Path(temp_dir)
            snapshot = work / args.source.name
            shutil.copyfile(args.source, snapshot)
            office = _run_libreoffice_pdf(
                snapshot,
                work,
                timeout,
                [args.source, output],
            )
            page_count = _validate_pdf_output(office.generated)
            selected = parse_render_pages(args.pages, page_count)
            try:
                document = pdfium.PdfDocument(str(office.generated))
            except Exception as exc:
                raise ToolError(
                    "external_tool_failure",
                    "PDFium could not open the converted PDF.",
                    details={"reason": str(exc)},
                ) from exc
            total_pixels = 0
            try:
                for page_number in selected:
                    page = document[page_number - 1]
                    width_pt, height_pt = page.get_size()
                    pixels = math.ceil(width_pt / 72 * dpi) * math.ceil(height_pt / 72 * dpi)
                    if pixels > MAX_RENDER_PIXELS_PER_PAGE:
                        raise ToolError(
                            "resource_limit",
                            "Rendered pixel count for one page exceeds the limit.",
                            details={
                                "page": page_number,
                                "pixels": pixels,
                                "limit": MAX_RENDER_PIXELS_PER_PAGE,
                            },
                        )
                    total_pixels += pixels
                    if total_pixels > MAX_RENDER_TOTAL_PIXELS:
                        raise ToolError(
                            "resource_limit",
                            "Total rendered pixel count exceeds the limit.",
                            details={"pixels": total_pixels, "limit": MAX_RENDER_TOTAL_PIXELS},
                        )
                    file_name = f"page-{page_number:03d}.png"
                    staged_file = staged / file_name
                    try:
                        page.render(scale=dpi / 72).to_pil().save(staged_file, format="PNG")
                    except Exception as exc:
                        raise ToolError(
                            "external_tool_failure",
                            "PDFium could not rasterize a PDF page.",
                            details={"page": page_number},
                        ) from exc
                    with staged_file.open("rb") as rendered:
                        if rendered.read(8) != b"\x89PNG\r\n\x1a\n":
                            raise ToolError(
                                "post_write_validation",
                                "Rendered file does not have a PNG signature.",
                                details={"file": file_name},
                            )
                    with Image.open(staged_file) as verified:
                        verified.load()
                        size = verified.size
                    if size[0] * size[1] > MAX_RENDER_PIXELS_PER_PAGE:
                        raise ToolError(
                            "post_write_validation",
                            "Rendered image exceeds the per-page pixel limit.",
                            details={"file": file_name},
                        )
                    data = staged_file.read_bytes()
                    images.append(
                        {
                            "page": page_number,
                            "file": file_name,
                            "width_px": size[0],
                            "height_px": size[1],
                            "bytes": len(data),
                            "sha256": hashlib.sha256(data).hexdigest(),
                        }
                    )
            finally:
                document.close()
        if sha256_file(args.source) != source_hash:
            fail(
                "post_write_validation",
                "Source changed before the destination publication gate.",
                path=args.source,
            )
        publish_directory_atomic(staged, output)
    except Exception:
        shutil.rmtree(staged, ignore_errors=True)
        raise
    reopened = 0
    for record in images:
        published = output / record["file"]
        data = published.read_bytes()
        if hashlib.sha256(data).hexdigest() != record["sha256"]:
            fail(
                "post_write_validation",
                "Published rendering does not match its staged hash.",
                file=record["file"],
            )
        with Image.open(published) as verified:
            verified.load()
        reopened += 1
    warnings = print_semantics_warnings(context, "Rendered page")
    if office.stderr.strip():
        warnings.append("LibreOffice emitted diagnostics; inspect conversion.diagnostics.")
    warnings.append(
        "Rendering exists to be looked at: open the published PNG pages and review layout "
        "before claiming visual correctness."
    )
    return success(
        "render",
        {
            "source": str(args.source.resolve()),
            "output": str(output.resolve()),
            "counts": {
                "sheets": len(context["sheet_names"]),
                "hidden_sheets": len(context["hidden_sheets"]),
                "pdf_pages": page_count,
                "pages_rendered": len(images),
            },
            "images": images,
            "fonts": context["fonts"],
            "conversion": {
                "engine": office.engine,
                "libreoffice": office.office_version,
                "timeout_seconds": timeout,
                "dpi": dpi,
                "diagnostics": {"stdout": office.stdout, "stderr": office.stderr},
            },
            "source_package": package,
            "verification": {
                "atomic_publish": True,
                "pdf_page_count": page_count,
                "png_reopened": reopened,
                "source_unchanged_at_publish_gate": True,
            },
        },
        warnings,
    )


def add_overwrite_argument(parser: argparse.ArgumentParser) -> None:
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="replace an existing destination, or permit source-equals-destination",
    )


def add_delimited_options(parser: argparse.ArgumentParser) -> None:
    parser.add_argument("--encoding", default="utf-8")
    parser.add_argument("--delimiter", help="single-character delimiter override")
    parser.add_argument(
        "--quoting",
        choices=["minimal", "all", "nonnumeric", "none"],
        default="minimal",
    )
    parser.add_argument("--na-value", default="")
    parser.add_argument("--date-format")
    parser.add_argument(
        "--allow-formulas",
        action="store_true",
        help="allow executable formula text instead of applying the injection guard",
    )


def build_parser() -> argparse.ArgumentParser:
    parser = JsonArgumentParser(description=__doc__)
    parser.add_argument(
        "--debug",
        action="store_true",
        help="include a traceback in unexpected JSON diagnostics",
    )
    subparsers = parser.add_subparsers(dest="operation", required=True)

    inspect_parser = subparsers.add_parser("inspect", help="inventory an XLSX workbook")
    inspect_parser.add_argument("source", type=Path)
    inspect_parser.add_argument("--sheet")
    inspect_parser.add_argument("--range")
    inspect_parser.add_argument("--no-cells", action="store_true")
    inspect_parser.add_argument("--max-cells", type=int, default=10_000)
    inspect_parser.set_defaults(handler=handle_inspect)

    extract_parser = subparsers.add_parser(
        "extract",
        help="extract a selected sheet or range to JSON, CSV, or TSV",
    )
    extract_parser.add_argument("source", type=Path)
    extract_parser.add_argument("output", type=Path)
    extract_parser.add_argument("--sheet")
    extract_parser.add_argument("--range")
    extract_parser.add_argument("--format", choices=["json", "csv", "tsv"], required=True)
    extract_parser.add_argument("--values", choices=["formulas", "cached"], default="formulas")
    add_delimited_options(extract_parser)
    add_overwrite_argument(extract_parser)
    extract_parser.set_defaults(handler=handle_extract)

    create_parser = subparsers.add_parser("create", help="create XLSX from a versioned job")
    create_parser.add_argument("job", type=Path)
    create_parser.add_argument("output", type=Path)
    create_parser.add_argument(
        "--cell-limit",
        type=int,
        default=MAX_DATA_CELLS,
        help="lower the cumulative job cell limit (maximum 2,000,000)",
    )
    add_overwrite_argument(create_parser)
    create_parser.set_defaults(handler=handle_create)

    edit_parser = subparsers.add_parser("edit", help="edit XLSX from a versioned job")
    edit_parser.add_argument("source", type=Path)
    edit_parser.add_argument("job", type=Path)
    edit_parser.add_argument("output", type=Path)
    edit_parser.add_argument(
        "--cell-limit",
        type=int,
        default=MAX_DATA_CELLS,
        help="lower the cumulative job cell limit (maximum 2,000,000)",
    )
    add_overwrite_argument(edit_parser)
    edit_parser.set_defaults(handler=handle_edit)

    clean_parser = subparsers.add_parser(
        "clean",
        help="apply explicit pandas cleanup policies into a separate sheet",
    )
    clean_parser.add_argument("source", type=Path)
    clean_parser.add_argument("job", type=Path)
    clean_parser.add_argument("output", type=Path)
    add_overwrite_argument(clean_parser)
    clean_parser.set_defaults(handler=handle_clean)

    summarize_parser = subparsers.add_parser(
        "summarize",
        help="add a static grouped/pivot-style summary sheet",
    )
    summarize_parser.add_argument("source", type=Path)
    summarize_parser.add_argument("job", type=Path)
    summarize_parser.add_argument("output", type=Path)
    add_overwrite_argument(summarize_parser)
    summarize_parser.set_defaults(handler=handle_summarize)

    convert_parser = subparsers.add_parser(
        "convert",
        help="convert XLSX to/from CSV or TSV, or export XLSX to PDF",
    )
    convert_parser.add_argument("source", type=Path)
    convert_parser.add_argument("output", type=Path)
    convert_parser.add_argument(
        "--input-format",
        choices=["auto", "xlsx", "csv", "tsv"],
        default="auto",
    )
    convert_parser.add_argument(
        "--output-format",
        choices=["auto", "xlsx", "csv", "tsv", "pdf"],
        default="auto",
    )
    convert_parser.add_argument(
        "--timeout",
        type=float,
        default=DEFAULT_EXTERNAL_TOOL_TIMEOUT,
        help="LibreOffice conversion deadline in seconds; PDF output only",
    )
    convert_parser.add_argument("--sheet")
    convert_parser.add_argument("--range")
    convert_parser.add_argument("--header-row", type=int)
    convert_parser.add_argument(
        "--sheet-policy",
        choices=["raw", "header"],
        default=None,
        help="XLSX export policy: raw used-range rows or header-aware rectangular data",
    )
    convert_parser.add_argument("--all-sheets", action="store_true")
    convert_parser.add_argument("--values", choices=["formulas", "cached"], default="cached")
    convert_parser.add_argument("--schema", type=Path)
    convert_parser.add_argument(
        "--leading-zeros",
        choices=["preserve", "numeric"],
        default=None,
    )
    convert_parser.add_argument("--index", action="store_true")
    convert_parser.add_argument("--no-header", action="store_true")
    convert_parser.add_argument(
        "--input-na-policy",
        choices=["literal", "default", "custom"],
        default="literal",
        help="CSV/TSV input NA recognition policy",
    )
    convert_parser.add_argument(
        "--input-na-value",
        action="append",
        default=[],
        help="custom CSV/TSV input NA marker; repeat with --input-na-policy custom",
    )
    convert_parser.add_argument(
        "--malformed-rows",
        choices=["error", "warn", "skip"],
        default="error",
    )
    add_delimited_options(convert_parser)
    add_overwrite_argument(convert_parser)
    convert_parser.set_defaults(handler=handle_convert)

    render_parser = subparsers.add_parser(
        "render",
        help="render workbook pages to PNG images via LibreOffice and PDFium",
    )
    render_parser.add_argument("source", type=Path)
    render_parser.add_argument("output", type=Path)
    render_parser.add_argument(
        "--dpi",
        type=int,
        default=96,
        help=f"raster resolution between {MIN_RENDER_DPI} and {MAX_RENDER_DPI}",
    )
    render_parser.add_argument(
        "--pages",
        help="comma-separated 1-based PDF page numbers; default renders every page",
    )
    render_parser.add_argument(
        "--timeout",
        type=float,
        default=DEFAULT_EXTERNAL_TOOL_TIMEOUT,
        help="LibreOffice conversion deadline in seconds",
    )
    render_parser.set_defaults(handler=handle_render)
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    parser = build_parser()
    debug = False
    try:
        args = parser.parse_args(argv)
        debug = args.debug
        ensure_dependencies()
        payload = args.handler(args)
        emit(payload)
        return 0
    except ToolError as exc:
        emit(
            {
                "schema_version": SCHEMA_VERSION,
                "ok": False,
                "error": {
                    "category": exc.category,
                    "message": str(exc),
                    "details": exc.details,
                },
            },
            stream=sys.stderr,
        )
        return exc.status
    except KeyboardInterrupt:
        emit(
            {
                "schema_version": SCHEMA_VERSION,
                "ok": False,
                "error": {
                    "category": "internal_error",
                    "message": "Operation was interrupted.",
                    "details": {},
                },
            },
            stream=sys.stderr,
        )
        return FAILURE_STATUS["internal_error"]
    except (LookupError, OverflowError, TypeError, ValueError) as exc:
        details: dict[str, Any] = {
            "exception_type": type(exc).__name__,
            "reason": str(exc),
        }
        if debug:
            details["traceback"] = traceback.format_exc()
        emit(
            {
                "schema_version": SCHEMA_VERSION,
                "ok": False,
                "error": {
                    "category": "bad_input",
                    "message": "A supplied value is invalid for the requested operation.",
                    "details": details,
                },
            },
            stream=sys.stderr,
        )
        return FAILURE_STATUS["bad_input"]
    except Exception as exc:  # pragma: no cover - defensive CLI boundary
        details: dict[str, Any] = {"exception_type": type(exc).__name__}
        if debug:
            details["traceback"] = traceback.format_exc()
        emit(
            {
                "schema_version": SCHEMA_VERSION,
                "ok": False,
                "error": {
                    "category": "internal_error",
                    "message": "Unexpected internal failure.",
                    "details": details,
                },
            },
            stream=sys.stderr,
        )
        return FAILURE_STATUS["internal_error"]


if __name__ == "__main__":
    raise SystemExit(main())
